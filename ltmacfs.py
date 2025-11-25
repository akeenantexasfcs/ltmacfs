import streamlit as st
import pandas as pd
import json
import io
import re
import os
import time
import hashlib
import tempfile
import shutil
import traceback
from datetime import datetime
from snowflake.snowpark.context import get_active_session
from fuzzywuzzy import fuzz, process
from openpyxl import load_workbook
# PDF Preview imports
try:
    from pdf2image import convert_from_bytes
    from PIL import Image
    pdf_preview_available = True
except ImportError:
    pdf_preview_available = False
# Test boto3 import
try:
    import boto3
    import botocore
    boto3_available = True
except ImportError as e:
    boto3_available = False
# Page config
st.set_page_config(
    page_title="LTMA - Cash Flow Statement Processing",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)
# Custom CSS for better styling
st.markdown("""
<style>
/* Main header styling */
.main-header {
    padding: 1rem;
    background: linear-gradient(90deg, #1e3c72 0%, #2a5298 100%);
    border-radius: 10px;
    margin-bottom: 2rem;
    color: white;
}
/* Section headers */
.section-header {
    padding: 0.5rem 1rem;
    background-color: #f0f2f6;
    border-left: 4px solid #1e3c72;
    margin: 1rem 0;
    border-radius: 0 5px 5px 0;
}
/* Workflow indicator */
.workflow-step {
    text-align: center;
    padding: 0.5rem;
    border-radius: 5px;
    font-weight: bold;
}
.workflow-active {
    background-color: #1e3c72;
    color: white;
}
.workflow-inactive {
    background-color: #e0e0e0;
    color: #666;
}
/* Custom button styling */
.stButton > button {
    width: 100%;
}
/* Metrics styling */
[data-testid="metric-container"] {
    background-color: #f0f2f6;
    border: 1px solid #ddd;
    padding: 1rem;
    border-radius: 5px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
}
/* Expander styling */
.streamlit-expanderHeader {
    background-color: #f8f9fa;
    border-radius: 5px;
}
/* Tab styling */
.stTabs [data-baseweb="tab-list"] {
    gap: 8px;
}
.stTabs [data-baseweb="tab"] {
    height: 50px;
    padding-left: 20px;
    padding-right: 20px;
    background-color: #f0f2f6;
    border-radius: 5px 5px 0 0;
}
.stTabs [aria-selected="true"] {
    background-color: #1e3c72;
    color: white;
}
/* Navigation styling */
.nav-section {
    background-color: #f8f9fa;
    padding: 1rem;
    border-radius: 8px;
    margin-bottom: 1rem;
}
.nav-header {
    font-weight: bold;
    color: #1e3c72;
    margin-bottom: 0.5rem;
    font-size: 1.1rem;
}
.nav-link {
    display: block;
    padding: 0.5rem 1rem;
    margin: 0.25rem 0;
    background-color: white;
    border-radius: 5px;
    color: #333;
    text-decoration: none;
    transition: all 0.3s ease;
}
.nav-link:hover {
    background-color: #e8f0fe;
    transform: translateX(5px);
}
.current-app {
    background-color: #1e3c72;
    color: white;
    padding: 0.5rem 1rem;
    border-radius: 5px;
    margin-top: 1rem;
}
/* Workflow diagram styling */
.workflow-container {
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 2rem;
    background-color: #f8f9fa;
    border-radius: 10px;
    margin: 2rem 0;
    flex-wrap: wrap;
}
.workflow-box {
    background-color: white;
    border: 2px solid #1e3c72;
    border-radius: 10px;
    padding: 1rem 2rem;
    margin: 0.5rem;
    text-align: center;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    transition: all 0.3s ease;
}
.workflow-box:hover {
    transform: translateY(-5px);
    box-shadow: 0 4px 8px rgba(0,0,0,0.2);
}
.workflow-arrow {
    font-size: 2rem;
    color: #1e3c72;
    margin: 0 1rem;
}
.overview-card {
    background-color: white;
    border-radius: 10px;
    padding: 2rem;
    margin: 1rem 0;
    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
}
.feature-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
    gap: 1.5rem;
    margin: 2rem 0;
}
.feature-card {
    background-color: #f8f9fa;
    border-radius: 8px;
    padding: 1.5rem;
    text-align: center;
    transition: all 0.3s ease;
}
.feature-card:hover {
    background-color: #e8f0fe;
    transform: translateY(-5px);
    box-shadow: 0 4px 12px rgba(0,0,0,0.15);
}
.feature-icon {
    font-size: 3rem;
    margin-bottom: 1rem;
}
/* PDF Preview specific styling */
.pdf-preview-container {
    border: 2px solid #e0e0e0;
    border-radius: 8px;
    padding: 1rem;
    background-color: #fafafa;
    margin: 1rem 0;
}
.page-selector {
    background-color: white;
    border: 1px solid #ddd;
    border-radius: 5px;
    padding: 0.5rem;
    margin: 0.5rem 0;
}
</style>
""", unsafe_allow_html=True)
# Initialize session state
if 'app_mode' not in st.session_state:
    st.session_state.app_mode = 'overview'
if 'processed_files' not in st.session_state:
    st.session_state.processed_files = {}
if 'aws_configured' not in st.session_state:
    st.session_state.aws_configured = False
if 'auto_update_dict' not in st.session_state:
    st.session_state.auto_update_dict = True
if 'cfs_dictionary_data' not in st.session_state:
    st.session_state.cfs_dictionary_data = None
if 'pdf_pages' not in st.session_state:
    st.session_state.pdf_pages = {}
if 'selected_pages' not in st.session_state:
    st.session_state.selected_pages = {}
# New: Cash Flow Workflow states
if 'cfs_json_files' not in st.session_state:
    st.session_state.cfs_json_files = {} # Stores all JSONs (uploaded or pushed)
if 'cfs_classifications' not in st.session_state:
    st.session_state.cfs_classifications = {} # Stores classification results
if 'cfs_column_names' not in st.session_state:
    st.session_state.cfs_column_names = {} # Stores column names for each JSON
if 'cfs_selected_tables' not in st.session_state:
    st.session_state.cfs_selected_tables = {} # Which tables selected (list of indices) per file
if 'cfs_processed_data' not in st.session_state:
    st.session_state.cfs_processed_data = {} # Final processed data ready for next step
if 'cfs_account_column' not in st.session_state:
    st.session_state.cfs_account_column = {} # Stores account column selection per file
if 'cfs_units_conversion' not in st.session_state:
    st.session_state.cfs_units_conversion = {} # Stores units conversion selection per file
if 'cfs_ready_for_aggregation' not in st.session_state:
    st.session_state.cfs_ready_for_aggregation = set() # Tracks which files are marked ready
if 'cfs_aggregated_data' not in st.session_state:
    st.session_state.cfs_aggregated_data = None
if 'show_removed_rows_message' not in st.session_state:
    st.session_state.show_removed_rows_message = False
if 'removed_zero_rows' not in st.session_state:
    st.session_state.removed_zero_rows = []
if 'cfs_user_selections' not in st.session_state:
    st.session_state.cfs_user_selections = {}
if 'cfs_mapping_results' not in st.session_state:
    st.session_state.cfs_mapping_results = None
if 'cfs_final_mapping_results' not in st.session_state:
    st.session_state.cfs_final_mapping_results = None
if 'cfs_expand_state' not in st.session_state:
    st.session_state.cfs_expand_state = {}
if 'cfs_all_labels' not in st.session_state:
    st.session_state.cfs_all_labels = []
if 'company_name' not in st.session_state:
    st.session_state.company_name = ""
if 'just_added_mappings' not in st.session_state:
    st.session_state.just_added_mappings = []
if 'show_success_message' not in st.session_state:
    st.session_state.show_success_message = False
if 'show_mapping' not in st.session_state:
    st.session_state.show_mapping = False
if 'dict_view_mode' not in st.session_state:
    st.session_state.dict_view_mode = "Full View"
if 'dict_search_term' not in st.session_state:
    st.session_state.dict_search_term = ""
if 'dict_selected_label' not in st.session_state:
    st.session_state.dict_selected_label = "All"
if 'dict_page_number' not in st.session_state:
    st.session_state.dict_page_number = 1
if 'workflow_step' not in st.session_state:
    st.session_state.workflow_step = "📋 Cash Flow Prep & AI Classification"
if 'cfs_selected_value_cols' not in st.session_state:
    st.session_state.cfs_selected_value_cols = {}
if 'cfs_prep_expander_state' not in st.session_state:
    st.session_state.cfs_prep_expander_state = {}
@st.cache_data
def get_stable_cache_key():
    """Generate a stable cache key for the current session."""
    return hashlib.md5(str(time.time()).encode()).hexdigest()[:8]
@st.cache_data(ttl=600)
def load_dictionary_cached(cache_key):
    """Load dictionary data with caching"""
    session = get_active_session()
    try:
        dict_data = session.sql("""
            SELECT ACCOUNT, LABEL, MNEMONIC, REFERENCE
            FROM LTMA.PUBLIC.CASH_FLOW_STATEMENT_COMPLEX_DICTIONARY
            ORDER BY LABEL, ACCOUNT
        """).to_pandas()
        return dict_data
    except Exception as e:
        st.error(f"Error loading dictionary: {str(e)}")
        return pd.DataFrame()
@st.cache_data(ttl=300)
def filter_dictionary_by_label(dictionary_df, label):
    """Cache filtered dictionary results for performance."""
    if dictionary_df is None or 'LABEL' not in dictionary_df.columns or label is None:
        return pd.DataFrame()
    return dictionary_df[dictionary_df['LABEL'].str.strip().str.lower() == label.strip().lower()].copy()
@st.cache_data
def create_summary_df(mapping_results, dictionary_df):
    summary_data = []
    label_account_sets = {}
    for label in dictionary_df['LABEL'].unique():
        label_account_sets[label] = set(dictionary_df[dictionary_df['LABEL'] == label]['ACCOUNT'].values)
    for result in mapping_results:
        label = result['label']
        account = result['account']
        is_new_flag = '🆕' if account not in label_account_sets.get(label, set()) else ''
        summary_data.append({
            'Label': label, 'Account': account,
            'Fuzzy Match': result['matches']['fuzzy']['match'] if result['matches']['fuzzy'] else 'N/A',
            'Fuzzy Score': f"{result['matches']['fuzzy']['score']}%" if result['matches']['fuzzy'] else '-',
            'AI Match': result['matches']['llm']['match'] if result['matches']['llm'] and 'error' not in result['matches']['llm'] else 'N/A',
            'New': is_new_flag
        })
    return pd.DataFrame(summary_data)
def clean_numeric_value(value):
    """Clean and convert value to numeric, handling negative numbers in parentheses.
    Returns a float value, defaulting to 0 for invalid inputs."""
    try:
        if pd.isna(value) or value == '' or value == '-':
            return 0
        value_str = str(value).strip()
        is_negative = False
        if value_str.startswith('(') and value_str.endswith(')'):
            is_negative = True
            value_str = value_str.lstrip('$(').rstrip(')')
        cleaned_value = re.sub(r'[$,]', '', value_str)
        result = float(cleaned_value)
        if is_negative:
            result *= -1
        return result
    except (ValueError, TypeError, AttributeError) as e:
        return 0
@st.cache_data
def aggregate_data(df):
    """Aggregate data by Label and Account, summing numeric columns.
    Returns a sorted dataframe with rounded values."""
    df = df.copy()
    df['Label'] = df['Label'].astype(str).str.strip()
    df['Account'] = df['Account'].astype(str).str.strip()
    numeric_cols = [col for col in df.columns if col not in ['Label', 'Account']]
    for col in numeric_cols:
        df[col] = df[col].apply(clean_numeric_value)
    aggregated = df.groupby(['Label', 'Account'], as_index=False)[numeric_cols].sum()
    for col in numeric_cols:
        aggregated[col] = aggregated[col].round(2)
    return aggregated
def sort_by_label_and_account(df, sort_column="Account"):
    """Sort dataframe by predefined label order and a specified column.
    Maintains cash flow structure: Operations, Investing, Financing, Other, Subtotal."""
    label_order = {
        "Cash from Operations": 0,
        "Cash from Investing": 1,
        "Cash from Financing": 2,
        "Cash from other": 3,
        "Subtotal": 4
    }
    if 'Label' in df.columns and sort_column in df.columns:
        df['_sort_order'] = df['Label'].map(label_order).fillna(5)
        df = df.sort_values(['_sort_order', sort_column]).drop('_sort_order', axis=1)
    return df.reset_index(drop=True)
def check_all_zeroes(df):
    """Check for rows where all numeric values are zero.
    Returns a boolean series indicating zero-value rows."""
    numeric_cols = [col for col in df.columns if col not in ['Label', 'Account']]
    if not numeric_cols:
        return pd.Series([False] * len(df))
    df_copy = df[numeric_cols].copy()
    df_copy = df_copy.fillna(0)
    return (df_copy == 0).all(axis=1)
@st.cache_data
def extract_tables_from_textract(data):
    """Extract tables from Textract JSON response.
    Returns a list of dataframes representing each table found."""
    tables = []
    for block in data['Blocks']:
        if block['BlockType'] == 'TABLE':
            table = {}
            if 'Relationships' in block:
                for relationship in block['Relationships']:
                    if relationship['Type'] == 'CHILD':
                        for cell_id in relationship['Ids']:
                            cell_block = next((b for b in data['Blocks'] if b['Id'] == cell_id), None)
                            if cell_block:
                                row_index = cell_block.get('RowIndex', 0)
                                col_index = cell_block.get('ColumnIndex', 0)
                                if row_index not in table:
                                    table[row_index] = {}
                                cell_text = ''
                                if 'Relationships' in cell_block:
                                    for rel in cell_block['Relationships']:
                                        if rel['Type'] == 'CHILD':
                                            for word_id in rel['Ids']:
                                                word_block = next((w for w in data['Blocks'] if w['Id'] == word_id), None)
                                                if word_block and word_block['BlockType'] == 'WORD':
                                                    cell_text += ' ' + word_block.get('Text', '')
                                table[row_index][col_index] = cell_text.strip()
            if table:
                table_df = pd.DataFrame.from_dict(table, orient='index').sort_index()
                table_df = table_df.sort_index(axis=1)
                tables.append(table_df)
    return tables
@st.cache_data(ttl=300) # Cache AI classifications for 5 minutes
def classify_cash_flow_with_ai_cached(accounts_json, model, label_list):
    """Cached version of AI classification to avoid repeated API calls"""
    session = get_active_session()
    prompt = f"""You are a financial analyst expert in cash flow statement classification.
Classify each cash flow statement line item into one of these categories:
* Cash from Operations
* Cash from Investing
* Cash from Financing
* Cash from other
* Subtotal
CRITICAL RULES FOR SUBTOTALS:
The following items MUST be classified as "Subtotal":
* "Net cash provided by operating activities" or similar
* "Net cash used in investing activities" or similar
* "Net cash provided by financing activities" or similar
* Any line item that starts with "Net cash" or "Total cash"
IMPORTANT: Any line item that starts with "Net" or "Total" should be classified as "Subtotal", NOT as Skip or any other category.
OTHER RULES:
* Only use "Skip" for section headers ending with ":", cash balance items like "Cash at beginning of period", "Cash at end of period", blank lines, or irrelevant headers.
* All individual operating items go in Cash from Operations
* All individual investing items go in Cash from Investing
* All individual financing items go in Cash from Financing
* Exchange rate effects or miscellaneous go in Cash from other
NEVER classify any "Net" or "Total" line as Skip!
Accounts to classify: {accounts_json}
Return a JSON array with the classification for each account in the same order.
Format: [{{"account": "account name", "category": "category name", "confidence": 0.95}}, ...]
Return ONLY the JSON array, no other text."""
    try:
        escaped_prompt = prompt.replace("'", "''")
        query = f"""
            SELECT SNOWFLAKE.CORTEX.COMPLETE(
                '{model}',
                '{escaped_prompt}'
            ) as response
        """
        result = session.sql(query).collect()
        if result and result[0]['RESPONSE']:
            response_text = result[0]['RESPONSE']
            try:
                json_start = response_text.find('[')
                json_end = response_text.rfind(']') + 1
                if json_start >= 0 and json_end > json_start:
                    json_str = response_text[json_start:json_end]
                    ai_results = json.loads(json_str)
                    return ai_results
                else:
                    return None
            except json.JSONDecodeError:
                return None
    except Exception:
        return None
def classify_cash_flow_with_ai(df, account_column, model, session):
    """Use AI to classify cash flow statement line items into standard categories.
    Returns a dictionary mapping dataframe indices to classifications with confidence scores."""
    classifications = {}
    accounts_list = []
    for idx, row in df.iterrows():
        account_name = str(row[account_column]).strip()
        if account_name and account_name not in ['', 'nan', 'None']:
            accounts_list.append({
                'index': idx,
                'account': account_name
            })
    if not accounts_list:
        st.warning("No valid accounts found for classification")
        return classifications
    accounts_json = json.dumps([acc['account'] for acc in accounts_list])
    # Use cached version
    ai_results = classify_cash_flow_with_ai_cached(accounts_json, model, tuple(accounts_list))
    if ai_results:
        for i, account_data in enumerate(accounts_list):
            if i < len(ai_results):
                classifications[account_data['index']] = {
                    'category': ai_results[i].get('category', 'Skip'),
                    'confidence': ai_results[i].get('confidence', 0.5)
                }
    else:
        # Fallback to rule-based classification
        classifications = rule_based_classification(accounts_list)
    return classifications
def rule_based_classification(accounts_list):
    """Fallback classification using rule-based logic when AI is unavailable.
    Returns classifications based on keyword matching."""
    classifications = {}
    for account_data in accounts_list:
        account = account_data['account'].lower()
        idx = account_data['index']
        if account.endswith(':') or account in ['operating activities', 'investing activities', 'financing activities'] or 'cash at beginning' in account or 'cash at end' in account or not account.strip():
            category = 'Skip'
        elif any(subtotal in account for subtotal in ['net cash', 'total cash']) and 'exchange' not in account:
            category = 'Subtotal'
        elif any(keyword in account for keyword in ['net income', 'depreciation', 'amortization', 'accounts receivable', 'operating']) and 'net cash' not in account:
            category = 'Cash from Operations'
        elif any(keyword in account for keyword in ['capital expenditure', 'acquisition', 'investment', 'property plant']) and 'net cash' not in account:
            category = 'Cash from Investing'
        elif any(keyword in account for keyword in ['debt', 'stock', 'dividend', 'financing']) and 'net cash' not in account:
            category = 'Cash from Financing'
        elif any(keyword in account for keyword in ['exchange rate', 'misc']) and 'net cash' not in account:
            category = 'Cash from other'
        else:
            category = 'Skip'
        classifications[idx] = {
            'category': category,
            'confidence': 0.7
        }
    return classifications
@st.cache_data(ttl=300) # Cache for 5 minutes
def get_all_matches(account_name, label, dictionary_df, cache_key, run_llm=False, model='llama3.1-70b', fuzzy_threshold=70):
    """Get both fuzzy and LLM matches for account mapping.
    Returns a dictionary with fuzzy and optional LLM match results."""
    results = {
        'fuzzy': None,
        'llm': None
    }
    # Use .copy() to avoid SettingWithCopyWarning later
    label_filtered = filter_dictionary_by_label(dictionary_df, label)
    if label_filtered.empty:
        return results
    # Ensure REFERENCE column exists for the logic below.
    if 'REFERENCE' not in label_filtered.columns:
        label_filtered['REFERENCE'] = label_filtered['ACCOUNT']
    # The list for matching should be the 'REFERENCE' variations.
    reference_list = label_filtered['REFERENCE'].dropna().tolist()
    if not reference_list:
        return results
    # Fuzzy matching
    fuzzy_matches = process.extract(account_name, reference_list, scorer=fuzz.token_set_ratio, limit=3)
    if fuzzy_matches and fuzzy_matches[0][1] > 0:
        best_fuzzy_match_name, best_fuzzy_score = fuzzy_matches[0]
        # Find the row in the dictionary that corresponds to the best matching REFERENCE
        match_row_df = label_filtered[label_filtered['REFERENCE'] == best_fuzzy_match_name]
        if not match_row_df.empty:
            match_row = match_row_df.iloc[0]
            results['fuzzy'] = {
                'match': best_fuzzy_match_name,
                'mnemonic': match_row['MNEMONIC'],
                'reference': match_row['REFERENCE'],
                'score': best_fuzzy_score
            }
    # LLM matching if requested
    if run_llm:
        try:
            session = get_active_session()
            sample_mappings = label_filtered.head(20)
            context = '\n'.join([f'"{row["ACCOUNT"]}" -> {row["MNEMONIC"]}' for _, row in sample_mappings.iterrows()])
            # The prompt should ask the LLM to choose from the REFERENCE list
            prompt = f"""You are a financial account mapping expert.
Find the best match for this account.
User Account: '{account_name}'
Label Category: '{label}'
Example mappings for {label}:
{context}
Available accounts to match (choose one):
{chr(10).join([f'"{ref}"' for ref in reference_list[:40]])}
Instructions:
1. Consider abbreviations (e.g., "A/R" = "Accounts Receivable")
2. Consider common variations (e.g., "Cash Flow from Operations" = "Net Cash Provided by Operating Activities")
3. Return ONLY the exact text from the available accounts list
Return the best matching account name exactly as shown above."""
            escaped_prompt = prompt.replace("'", "''")
            response = session.sql(f"""
                SELECT SNOWFLAKE.CORTEX.COMPLETE(
                    '{model}',
                    '{escaped_prompt}'
                ) as match
            """).collect()[0]['MATCH']
            llm_result = response.strip().strip('"').strip("'")
            llm_match = label_filtered[label_filtered['REFERENCE'] == llm_result]
            if not llm_match.empty:
                match_row = llm_match.iloc[0]
                results['llm'] = {
                    'match': llm_result,
                    'mnemonic': match_row['MNEMONIC'],
                    'reference': match_row['REFERENCE'],
                    'confidence': 'High'
                }
            else:
                # Case-insensitive fallback
                for _, row in label_filtered.iterrows():
                    if row['REFERENCE'].lower() == llm_result.lower():
                        results['llm'] = {
                            'match': row['REFERENCE'],
                            'mnemonic': row['MNEMONIC'],
                            'reference': row['REFERENCE'],
                            'confidence': 'Medium'
                        }
                        break
        except Exception as e:
            results['llm'] = {'error': str(e)}
    return results
# AWS Textract helper functions
def extract_tables_from_blocks(blocks):
    """Extract table data from Textract blocks.
    Returns a list of table data as 2D arrays."""
    tables = []
    blocks_map = {block.get('Id'): block for block in blocks}
    for block in blocks:
        if block.get('BlockType') == 'TABLE':
            table = extract_table_data(block, blocks_map)
            tables.append(table)
    return tables
def get_table_block_info(blocks):
    """Get information about which blocks belong to which table.
    Returns a list of dictionaries with table block and associated block IDs."""
    table_info = []
    blocks_map = {block.get('Id'): block for block in blocks}
    for block in blocks:
        if block.get('BlockType') == 'TABLE':
            table_block_ids = {block['Id']}
            for relationship in block.get('Relationships', []):
                if relationship.get('Type') == 'CHILD':
                    for cell_id in relationship.get('Ids', []):
                        table_block_ids.add(cell_id)
                        cell_block = blocks_map.get(cell_id, {})
                        for cell_rel in cell_block.get('Relationships', []):
                            if cell_rel.get('Type') == 'CHILD':
                                for word_id in cell_rel.get('Ids', []):
                                    table_block_ids.add(word_id)
            table_info.append({
                'table_block': block,
                'all_block_ids': table_block_ids
            })
    return table_info
def filter_blocks_by_selected_tables(blocks, selected_table_indices):
    """Filter blocks to only include those from selected tables.
    Returns filtered list of blocks."""
    if not selected_table_indices:
        return [b for b in blocks if b.get('BlockType') not in ['TABLE', 'CELL']]
    table_info = get_table_block_info(blocks)
    included_block_ids = set()
    for idx in selected_table_indices:
        if idx < len(table_info):
            included_block_ids.update(table_info[idx]['all_block_ids'])
    blocks_map = {block.get('Id'): block for block in blocks}
    all_table_block_ids = set()
    for info in table_info:
        all_table_block_ids.update(info['all_block_ids'])
    for block in blocks:
        if block.get('Id') not in all_table_block_ids:
            included_block_ids.add(block.get('Id'))
    filtered_blocks = [b for b in blocks if b.get('Id') in included_block_ids]
    return filtered_blocks
def extract_table_data(table_block, blocks_map):
    """Extract data from a single table block.
    Returns a 2D array representing table cells."""
    rows = []
    for relationship in table_block.get('Relationships', []):
        if relationship.get('Type') == 'CHILD':
            for cell_id in relationship.get('Ids', []):
                cell = blocks_map.get(cell_id, {})
                if cell.get('BlockType') == 'CELL':
                    row_index = cell.get('RowIndex', 1) - 1
                    col_index = cell.get('ColumnIndex', 1) - 1
                    while len(rows) <= row_index:
                        rows.append([])
                    while len(rows[row_index]) <= col_index:
                        rows[row_index].append('')
                    cell_text = ''
                    for cell_relationship in cell.get('Relationships', []):
                        if cell_relationship.get('Type') == 'CHILD':
                            for word_id in cell_relationship.get('Ids', []):
                                word = blocks_map.get(word_id, {})
                                if word.get('BlockType') == 'WORD':
                                    cell_text += word.get('Text', '') + ' '
                    rows[row_index][col_index] = cell_text.strip()
    return rows
def process_pdf_pages(file_bytes, textract, s3, cred_data, selected_pages):
    """Process selected pages from a PDF through AWS Textract.
    Returns all blocks from the document analysis."""
    all_blocks = []
    # Generate unique object name
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    object_name = f"textract-temp/{timestamp}_selected_pages.pdf"
    # Upload PDF to S3
    s3.put_object(
        Body=file_bytes,
        Bucket=cred_data['BUCKET_NAME'],
        Key=object_name
    )
    try:
        # Start document analysis with page selection
        response = textract.start_document_analysis(
            DocumentLocation={
                'S3Object': {
                    'Bucket': cred_data['BUCKET_NAME'],
                    'Name': object_name
                }
            },
            FeatureTypes=['TABLES']
        )
        job_id = response['JobId']
        attempt = 0
        # Wait for job to complete
        while attempt < 60:
            response = textract.get_document_analysis(JobId=job_id)
            job_status = response['JobStatus']
            if job_status == 'SUCCEEDED':
                # Get all blocks
                all_blocks_raw = response.get('Blocks', [])
                next_token = response.get('NextToken')
                while next_token:
                    response = textract.get_document_analysis(
                        JobId=job_id,
                        NextToken=next_token
                    )
                    all_blocks_raw.extend(response.get('Blocks', []))
                    next_token = response.get('NextToken')
                # Filter blocks by selected pages
                for block in all_blocks_raw:
                    page_num = block.get('Page', 1)
                    if page_num in selected_pages:
                        all_blocks.append(block)
                break
            elif job_status == 'FAILED':
                raise Exception("Textract analysis failed")
            time.sleep(2)
            attempt += 1
    finally:
        # Clean up S3 object
        try:
            s3.delete_object(Bucket=cred_data['BUCKET_NAME'], Key=object_name)
        except:
            pass
    return all_blocks
def process_single_file(file, textract, s3, cred_data, selected_pages=None):
    """Process a single file through AWS Textract.
    Returns all blocks from the document analysis."""
    file_bytes = file.read()
    file_extension = os.path.splitext(file.name)[1].lower()
    all_blocks = []
    if file_extension == '.pdf':
        if selected_pages:
            # Process only selected pages
            all_blocks = process_pdf_pages(file_bytes, textract, s3, cred_data, selected_pages)
        else:
            # Process entire PDF
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            object_name = f"textract-temp/{timestamp}_{file.name}"
            s3.put_object(
                Body=file_bytes,
                Bucket=cred_data['BUCKET_NAME'],
                Key=object_name
            )
            response = textract.start_document_analysis(
                DocumentLocation={
                    'S3Object': {
                        'Bucket': cred_data['BUCKET_NAME'],
                        'Name': object_name
                    }
                },
                FeatureTypes=['TABLES']
            )
            job_id = response['JobId']
            attempt = 0
            while attempt < 60:
                response = textract.get_document_analysis(JobId=job_id)
                job_status = response['JobStatus']
                if job_status == 'SUCCEEDED':
                    all_blocks.extend(response.get('Blocks', []))
                    next_token = response.get('NextToken')
                    while next_token:
                        response = textract.get_document_analysis(
                            JobId=job_id,
                            NextToken=next_token
                        )
                        all_blocks.extend(response.get('Blocks', []))
                        next_token = response.get('NextToken')
                    break
                elif job_status == 'FAILED':
                    raise Exception("Textract analysis failed")
                time.sleep(2)
                attempt += 1
            try:
                s3.delete_object(Bucket=cred_data['BUCKET_NAME'], Key=object_name)
            except:
                pass
    else:
        # Process images directly
        response = textract.analyze_document(
            Document={'Bytes': file_bytes},
            FeatureTypes=['TABLES']
        )
        all_blocks = response.get('Blocks', [])
    return all_blocks
def preview_pdf_pages(file):
    """Convert PDF to images for preview.
    Returns list of PIL images."""
    if not pdf_preview_available:
        return None
    try:
        file_bytes = file.read()
        # Reset file pointer for later use
        file.seek(0)
        # Convert PDF to images
        images = convert_from_bytes(file_bytes, dpi=150)
        return images
    except Exception as e:
        st.error(f"Error previewing PDF: {str(e)}")
        return None
# CIQ Template Processing Function (kept intact)
def process_ciq_template(template_type, session):
    """Process CIQ template with uploaded financial data files."""
    unique_id = template_type.lower()
    # Upload widgets
    uploaded_template = st.file_uploader(
        f"Upload CIQ Template ({template_type})",
        type=['xlsx', 'xlsm'],
        key=f"ciq_template_{unique_id}"
    )
    uploaded_balance_sheet = st.file_uploader(
        f"Upload Completed Balance Sheet Data ({template_type})",
        type=['xlsx', 'xlsm'],
        key=f"ciq_balance_sheet_{unique_id}"
    )
    uploaded_cash_flow = st.file_uploader(
        f"Upload Completed Cash Flow Statement ({template_type})",
        type=['xlsx', 'xlsm'],
        key=f"ciq_cash_flow_{unique_id}"
    )
    uploaded_income_statement = st.file_uploader(
        f"Upload Completed Income Statement Data ({template_type})",
        type=['xlsx', 'xlsm'],
        key=f"ciq_income_statement_{unique_id}"
    )
    if st.button(f"🚀 Populate {template_type} Template Now", key=f"ciq_populate_button_{unique_id}", type="primary", use_container_width=True):
        if not uploaded_template:
            st.error("Please upload a template file.")
            return
        if not (uploaded_balance_sheet or uploaded_cash_flow or uploaded_income_statement):
            st.error("Please upload at least one data file.")
            return
        try:
            temp_dir = tempfile.mkdtemp()
            template_path = os.path.join(temp_dir, "template.xlsx")
            with open(template_path, "wb") as f:
                f.write(uploaded_template.read())
            template_wb = load_workbook(template_path, keep_vba=True)
            def process_sheet(uploaded_file, sheet_name, row_range, date_row):
                if not uploaded_file:
                    return
                file_path = os.path.join(temp_dir, f"{sheet_name}.xlsx")
                with open(file_path, "wb") as f:
                    f.write(uploaded_file.read())
                try:
                    sheet_wb = load_workbook(file_path, keep_vba=True)
                    as_presented_sheet = sheet_wb[f"As Presented - {sheet_name}"]
                    standardized_sheet = pd.read_excel(file_path, sheet_name=f"Standardized - {sheet_name}")
                    if 'CIQ' not in standardized_sheet.columns:
                        st.error(f"The column 'CIQ' is missing from the Standardized - {sheet_name} sheet.")
                        return
                    st.success(f"✅ {sheet_name} CIQ column found, proceeding...")
                    # Copy "As Presented" sheet to template
                    if f"As Presented - {sheet_name}" in template_wb.sheetnames:
                        del template_wb[f"As Presented - {sheet_name}"]
                    new_sheet = template_wb.create_sheet(f"As Presented - {sheet_name}")
                    for row in as_presented_sheet.iter_rows():
                        for cell in row:
                            new_sheet[cell.coordinate].value = cell.value
                    new_sheet.sheet_properties.tabColor = "FFA500"
                    # Copy "Standardized" sheet to template
                    if f"Standardized - {sheet_name}" in template_wb.sheetnames:
                        del template_wb[f"Standardized - {sheet_name}"]
                    standardized_ws = template_wb.create_sheet(f"Standardized - {sheet_name}")
                    for col_num, header in enumerate(standardized_sheet.columns, 1):
                        standardized_ws.cell(row=1, column=col_num, value=header)
                    for r_idx, row in enumerate(standardized_sheet.itertuples(index=False), 2):
                        for c_idx, value in enumerate(row, 1):
                            standardized_ws.cell(row=r_idx, column=c_idx, value=value)
                    # Get values for Upload sheet
                    upload_sheet = template_wb["Upload"]
                    ciq_values = standardized_sheet['CIQ'].tolist()
                    dates = list(standardized_sheet.columns[2:])
                    # Update the upload sheet based on template type
                    if template_type == "Annual":
                        for row in upload_sheet.iter_rows(min_row=row_range[0], max_row=row_range[1], min_col=4, max_col=upload_sheet.max_column):
                            ciq_cell = upload_sheet.cell(row=row[0].row, column=11)
                            ciq_value = ciq_cell.value
                            if ciq_value in ciq_values:
                                for col in range(4, 10):
                                    date_value = upload_sheet.cell(row=date_row, column=col).value
                                    if date_value in dates:
                                        lookup_value = standardized_sheet.loc[standardized_sheet['CIQ'] == ciq_value, date_value].sum()
                                        if not pd.isna(lookup_value):
                                            cell_to_update = upload_sheet.cell(row=row[0].row, column=col)
                                            if cell_to_update.data_type == 'f' or cell_to_update.value is None:
                                                cell_to_update.value = lookup_value
                        # Process negation row
                        for row in upload_sheet.iter_rows(min_row=row_range[1] + 1, max_row=row_range[1] + 1, min_col=4, max_col=9):
                            for cell in row:
                                if cell.value is not None:
                                    try:
                                        cell_value = float(cell.value)
                                        cell.value = -abs(cell_value)
                                    except ValueError:
                                        pass
                    elif template_type == "Quarterly":
                        for row in upload_sheet.iter_rows(min_row=row_range[0], max_row=row_range[1], min_col=4, max_col=21):
                            ciq_cell = upload_sheet.cell(row=row[0].row, column=23)
                            ciq_value = ciq_cell.value
                            if ciq_value in ciq_values:
                                for col in range(4, 22):
                                    date_value = upload_sheet.cell(row=date_row, column=col).value
                                    if date_value in dates:
                                        lookup_value = standardized_sheet.loc[standardized_sheet['CIQ'] == ciq_value, date_value].sum()
                                        if not pd.isna(lookup_value):
                                            cell_to_update = upload_sheet.cell(row=row[0].row, column=col)
                                            if cell_to_update.data_type == 'f' or cell_to_update.value is None:
                                                cell_to_update.value = lookup_value
                        # Process negation row
                        for row in upload_sheet.iter_rows(min_row=row_range[1] + 1, max_row=row_range[1] + 1, min_col=4, max_col=21):
                            for cell in row:
                                if cell.value is not None:
                                    try:
                                        cell_value = float(cell.value)
                                        cell.value = -abs(cell_value)
                                    except ValueError:
                                        pass
                except Exception as e:
                    st.error(f"Error processing {sheet_name}: {str(e)}")
                    st.error(traceback.format_exc())
            # Process each uploaded file
            if template_type == "Annual":
                if uploaded_balance_sheet:
                    process_sheet(uploaded_balance_sheet, "Balance Sheet", (96, 162), 94)
                if uploaded_cash_flow:
                    process_sheet(uploaded_cash_flow, "Cash Flow", (171, 234), 169)
                if uploaded_income_statement:
                    process_sheet(uploaded_income_statement, "Income Stmt", (14, 72), 12)
            elif template_type == "Quarterly":
                if uploaded_balance_sheet:
                    process_sheet(uploaded_balance_sheet, "Balance Sheet", (96, 162), 94)
                if uploaded_cash_flow:
                    process_sheet(uploaded_cash_flow, "Cash Flow", (171, 234), 169)
                if uploaded_income_statement:
                    process_sheet(uploaded_income_statement, "Income Stmt", (14, 72), 12)
            # Save the updated workbook
            output_path = os.path.join(temp_dir, "output.xlsx")
            template_wb.save(output_path)
            # Read file for download
            with open(output_path, "rb") as f:
                output_data = f.read()
            # Provide download button
            st.download_button(
                label=f"📥 Download Updated {template_type} Template",
                data=output_data,
                file_name=f"Updated_{uploaded_template.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success(f"✅ {template_type} Template populated successfully!")
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
            st.error(traceback.format_exc())
        finally:
            # Clean up temp files
            try:
                shutil.rmtree(temp_dir)
            except:
                pass
# Main app
def main():
    # Get Snowflake session
    session = get_active_session()
    # AI Models
    MODELS = [
        "llama3.1-405b",
        "llama3.1-70b",
        "snowflake-llama-3.3-70b",
        "llama4-maverick",
        "claude-3-7-sonnet",
        "claude-4-sonnet"
    ]
    # Get stable cache key
    cache_key = get_stable_cache_key()
    # Sidebar navigation
    with st.sidebar:
        st.markdown("## 💼 LTMA Suite")
        st.markdown("#### Less Time More Analysis")
        st.markdown("---")
        # Navigation section
        st.markdown('<div class="nav-section">', unsafe_allow_html=True)
        st.markdown('<div class="nav-header">🧭 Navigation</div>', unsafe_allow_html=True)
        # Quick Links
        with st.expander("🔗 Quick Links", expanded=True):
            # Main Application Overview
            if st.button("🏠 Main Application", use_container_width=True, type="primary" if st.session_state.app_mode == 'overview' else "secondary", key="nav_main_app"):
                st.session_state.app_mode = 'overview'
                st.rerun()
            # Document Processing
            if st.button("📄 Document Processing", use_container_width=True, type="primary" if st.session_state.app_mode == 'textract' else "secondary", key="nav_doc_processing"):
                st.session_state.app_mode = 'textract'
                st.rerun()
            # Cash Flow Workflow
            if st.button("💰 Cash Flow Workflow", use_container_width=True, type="primary" if st.session_state.app_mode == 'main' else "secondary", key="nav_cash_flow"):
                st.session_state.app_mode = 'main'
                st.session_state.workflow_step = "📋 Cash Flow Prep & AI Classification"
                st.rerun()
            # Balance Sheet Processing
            st.page_link("https://app.snowflake.com/gubfsqd/owb27392/#/streamlit-apps/CORTEX_SEARCH_TUTORIAL_DB.PUBLIC.E9UOE6FO25O_E7VE", label="📊 Balance Sheet Processing", use_container_width=True)
            # Income Statement Processing
            st.page_link("https://app.snowflake.com/gubfsqd/owb27392/#/streamlit-apps/LTMA.PUBLIC.XR0S9H476U7WKO00", label="📈 Income Statement Processing", use_container_width=True)
            # CIQ Template Upload
            if st.button("📤 CIQ Template Upload", use_container_width=True, type="primary" if st.session_state.app_mode == 'ciq' else "secondary", key="nav_ciq_upload"):
                st.session_state.app_mode = 'ciq'
                st.rerun()
            # Data Dictionary
            if st.button("📚 Data Dictionary", use_container_width=True, type="secondary", key="nav_dictionary"):
                st.session_state.app_mode = 'main'
                st.session_state.workflow_step = "📚 Data Dictionary Management"
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
        st.markdown("---")
        # Current App indicator
        st.markdown('<div class="current-app">', unsafe_allow_html=True)
        if st.session_state.app_mode == 'overview':
            st.markdown("📍 Current App: LTMA Overview")
        elif st.session_state.app_mode == 'main':
            st.markdown("📍 Current App: Cash Flow Workflow")
        elif st.session_state.app_mode == 'textract':
            st.markdown("📍 Current App: Document Processing")
        elif st.session_state.app_mode == 'ciq':
            st.markdown("📍 Current App: CIQ Template Upload")
        st.markdown('</div>', unsafe_allow_html=True)
        st.markdown("---")
        # Settings section
        st.markdown("### ⚙️ Settings")
        # Only show relevant settings based on mode
        if st.session_state.app_mode == 'main':
            use_llm = st.checkbox("Enable AI Matching", value=True, help="Run AI analysis alongside fuzzy matching")
            model = MODELS[0] if use_llm else None
            # Dictionary update control
            st.markdown("#### 📚 Dictionary Settings")
            st.session_state.auto_update_dict = st.checkbox(
                "Enable Dictionary Updates",
                value=st.session_state.auto_update_dict,
                help="Allow adding new account mappings to the dictionary"
            )
            if st.session_state.auto_update_dict:
                st.info("New mappings will be identified and can be added to improve future matching")
            # UI Controls
            st.markdown("#### 🎛️ UI Controls")
            auto_collapse = st.checkbox("Auto-Collapse Sections", value=False, help="Automatically collapse completed sections")
            expand_all = st.button("📂 Expand All Sections", key="expand_all_btn")
            collapse_all = st.button("📁 Collapse All Sections", key="collapse_all_btn")
            if 'cfs_expand_state' not in st.session_state:
                st.session_state.cfs_expand_state = {}
            if expand_all:
                st.session_state.cfs_expand_state = {label: True for label in st.session_state.get('cfs_all_labels', [])}
                st.rerun()
            if collapse_all:
                st.session_state.cfs_expand_state = {label: False for label in st.session_state.get('cfs_all_labels', [])}
                st.rerun()
            st.markdown("#### 📖 Matching Methods")
            with st.expander("🔍 Fuzzy Matching", expanded=False):
                st.write("""
What it does: Compares text similarity between account names
How it works:
* Handles spelling variations (e.g., "A/R" ↔ "Accounts Receivable")
* Ignores word order (e.g., "Cash and equivalents" ↔ "Equivalents and cash")
* Tolerates minor differences
Score meaning:
* 100% = Exact match
* 80%+ = High confidence
* 70-79% = Medium confidence
* <70% = Low confidence
""")
            with st.expander("🤖 AI Matching", expanded=False):
                st.write("""
What it does: Uses LLM to understand financial context
How it works:
* Understands abbreviations
* Recognizes synonyms
* Applies accounting knowledge
Best for:
* Complex account names
* Industry-specific terms
* When fuzzy match fails
""")
            with st.expander("✏️ Manual Selection", expanded=False):
                st.write("""
What it does: Direct user selection
When to use:
* Neither method finds good match
* Company-specific accounts
* Special mapping requirements
""")
        elif st.session_state.app_mode == 'textract':
            # AWS Configuration display
            st.markdown("#### 📋 AWS Configuration")
            if st.session_state.aws_configured:
                st.success("✅ AWS Connected")
                if 'aws_creds' in st.session_state:
                    st.info(f"""
Configuration:
* Bucket: {st.session_state.aws_creds['BUCKET_NAME']}
* Region: {st.session_state.aws_creds['AWS_REGION']}
""")
            else:
                st.warning("⚠️ AWS not connected")
                st.info("Click 'Test AWS Connection' in the main panel to connect")
    # Main content area
    if st.session_state.app_mode == 'overview':
        # LTMA Overview Page
        st.markdown('<div class="main-header">🚀 LTMA - Less Time More Analysis<br>Streamline your financial statement processing and standardization</div>', unsafe_allow_html=True)
        # Introduction
        st.markdown('<div class="overview-card">', unsafe_allow_html=True)
        st.markdown("""
Welcome to LTMA Suite
LTMA is a comprehensive financial data processing platform designed to help analysts and finance professionals spend less time on data preparation and more time on analysis.
Our suite automates the tedious process of extracting, standardizing, and mapping financial statements to industry-standard formats, enabling you to focus on what matters most - insights and decision-making.
""")
        st.markdown('</div>', unsafe_allow_html=True)
        # Workflow Diagram
        st.markdown("### 📊 End-to-End Workflow")
        st.markdown('<div class="workflow-container">', unsafe_allow_html=True)
        st.markdown("""
📄<br>Document<br>Processing
→
📊<br>Balance Sheet<br>Workflow
→
📈<br>Income Statement<br>Workflow
→
💰<br>Cash Flow<br>Workflow
→
📤<br>CIQ Template<br>Upload
""", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        # Features Grid
        st.markdown("### 🌟 Key Features")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown('<div class="feature-card">', unsafe_allow_html=True)
            st.markdown("""
🤖<br>AI-Powered Classification
Leverage advanced AI models to automatically classify financial statement line items with high accuracy
""", unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        with col2:
            st.markdown('<div class="feature-card">', unsafe_allow_html=True)
            st.markdown("""
🔍<br>Smart Matching
Combine fuzzy matching and AI to map company-specific accounts to standard mnemonics
""", unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        with col3:
            st.markdown('<div class="feature-card">', unsafe_allow_html=True)
            st.markdown("""
📚<br>Learning Dictionary
Build and maintain a company-specific dictionary that improves matching accuracy over time
""", unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
    elif st.session_state.app_mode == 'textract':
        # Document Processing Mode
        st.markdown('<div class="main-header">📄 Document Processing Center<br>Extract tables from documents using AWS Textract with PDF preview and page selection</div>', unsafe_allow_html=True)
        # AWS Textract Component
        st.markdown('<div class="section-header">AWS Textract Integration</div>', unsafe_allow_html=True)
        # Configuration
        AWS_CREDS_DATABASE = "CORTEX_SEARCH_TUTORIAL_DB"
        AWS_CREDS_SCHEMA = "SECRETS"
        AWS_CREDS_TABLE = "AWS_CREDENTIALS"
        # Check boto3
        if not boto3_available:
            st.error("boto3 not available. Please check package configuration.")
            st.stop()
        # Check PDF preview libraries
        if not pdf_preview_available:
            st.warning("⚠️ PDF preview not available. Install pdf2image and Pillow for preview functionality.")
        # Load and test credentials
        try:
            # Get credentials
            cred_result = session.sql(f"""
                SELECT AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY, BUCKET_NAME, AWS_REGION
                FROM {AWS_CREDS_DATABASE}.{AWS_CREDS_SCHEMA}.{AWS_CREDS_TABLE}
                WHERE CREDENTIAL_NAME = 'DEFAULT'
            """).collect()
            # Convert Row object to dictionary
            if cred_result:
                cred_row = cred_result[0]
                cred_data = {
                    'AWS_ACCESS_KEY_ID': cred_row['AWS_ACCESS_KEY_ID'],
                    'AWS_SECRET_ACCESS_KEY': cred_row['AWS_SECRET_ACCESS_KEY'],
                    'BUCKET_NAME': cred_row['BUCKET_NAME'],
                    'AWS_REGION': cred_row['AWS_REGION']
                }
            else:
                st.error("No AWS credentials found in database")
                st.stop()
            # Test connection button
            if st.button("🔐 Test AWS Connection", use_container_width=True, key="test_aws_connection"):
                try:
                    # Create session
                    aws_session = boto3.Session(
                        aws_access_key_id=cred_data['AWS_ACCESS_KEY_ID'],
                        aws_secret_access_key=cred_data['AWS_SECRET_ACCESS_KEY'],
                        region_name=cred_data['AWS_REGION']
                    )
                    # Test STS
                    sts = aws_session.client('sts')
                    identity = sts.get_caller_identity()
                    st.success(f"""
✅ AWS Connection Successful!
* Account: {identity['Account']}
* User: {identity['Arn'].split('/')[-1]}
""")
                    # Store session
                    st.session_state.aws_configured = True
                    st.session_state.aws_session = aws_session
                    st.session_state.aws_creds = cred_data
                except Exception as e:
                    st.error(f"Connection failed: {str(e)}")
            # Main processing
            if st.session_state.aws_configured:
                st.divider()
                # Step 1: File upload - MULTIPLE FILES
                st.markdown("### 📤 Step 1: Upload Documents")
                uploaded_files = st.file_uploader(
                    "Upload PDF documents for table extraction:",
                    type=['pdf'],
                    accept_multiple_files=True,
                    key="pdf_uploader"
                )
                if uploaded_files:
                    st.success(f"✅ Uploaded {len(uploaded_files)} file(s)")
                # Step 2: PDF Preview and Page Selection for ALL files
                st.divider()
                st.markdown("### 👁️ Step 2: Preview and Select Pages from Each PDF")
                # Initialize dicts if needed
                if not isinstance(st.session_state.pdf_pages, dict):
                    st.session_state.pdf_pages = {}
                if not isinstance(st.session_state.selected_pages, dict):
                    st.session_state.selected_pages = {}
                # Process each uploaded file
                for file_idx, uploaded_file in enumerate(uploaded_files):
                    with st.expander(f"📄 {uploaded_file.name} ({uploaded_file.size / 1024:.1f} KB)", expanded=(file_idx == 0)):
                        # Generate preview if not already done
                        if uploaded_file.name not in st.session_state.pdf_pages:
                            if pdf_preview_available:
                                with st.spinner(f"Generating preview for {uploaded_file.name}..."):
                                    images = preview_pdf_pages(uploaded_file)
                                    if images:
                                        st.session_state.pdf_pages[uploaded_file.name] = images
                                        # Initialize all pages selected by default
                                        st.session_state.selected_pages[uploaded_file.name] = list(range(1, len(images) + 1))
                            else:
                                st.warning("PDF preview not available. Please specify page numbers manually.")
                                if uploaded_file.name not in st.session_state.selected_pages:
                                    st.session_state.selected_pages[uploaded_file.name] = []
                        # Display PDF preview and page selection
                        if pdf_preview_available and uploaded_file.name in st.session_state.pdf_pages:
                            images = st.session_state.pdf_pages[uploaded_file.name]
                            st.info(f"📄 Total pages: {len(images)}")
                            # Page selection controls
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                if st.button("☑️ Select All", key=f"select_all_{file_idx}"):
                                    st.session_state.selected_pages[uploaded_file.name] = list(range(1, len(images) + 1))
                                    st.rerun()
                            with col2:
                                if st.button("☐ Deselect All", key=f"deselect_all_{file_idx}"):
                                    st.session_state.selected_pages[uploaded_file.name] = []
                                    st.rerun()
                            with col3:
                                # Calculate current selected count
                                current_selected = st.session_state.selected_pages.get(uploaded_file.name, [])
                                st.metric("Selected Pages", len(current_selected))
                            st.markdown("---")
                            # Display pages in grid with checkboxes
                            cols_per_row = 3
                            for i in range(0, len(images), cols_per_row):
                                cols = st.columns(cols_per_row)
                                for j in range(cols_per_row):
                                    idx = i + j
                                    if idx < len(images):
                                        with cols[j]:
                                            page_num = idx + 1
                                            # Get current selection state
                                            current_selection = st.session_state.selected_pages.get(uploaded_file.name, [])
                                            is_selected = page_num in current_selection
                                            # Checkbox for page selection
                                            selected = st.checkbox(
                                                f"Page {page_num}",
                                                value=is_selected,
                                                key=f"page_{file_idx}_{page_num}"
                                            )
                                            # Update selection immediately
                                            if selected and not is_selected:
                                                if uploaded_file.name not in st.session_state.selected_pages:
                                                    st.session_state.selected_pages[uploaded_file.name] = []
                                                if page_num not in st.session_state.selected_pages[uploaded_file.name]:
                                                    st.session_state.selected_pages[uploaded_file.name].append(page_num)
                                                st.session_state.selected_pages[uploaded_file.name].sort()
                                            elif not selected and is_selected:
                                                if uploaded_file.name in st.session_state.selected_pages:
                                                    if page_num in st.session_state.selected_pages[uploaded_file.name]:
                                                        st.session_state.selected_pages[uploaded_file.name].remove(page_num)
                                            # Display page preview
                                            st.image(
                                                images[idx],
                                                use_container_width=True,
                                                caption=f"Page {page_num}"
                                            )
                        else:
                            # Manual page selection fallback
                            st.markdown("Enter page numbers to process (comma-separated):")
                            page_input = st.text_input(
                                "Example: 1,2,3 or 1-5",
                                value="",
                                key=f"manual_page_input_{file_idx}"
                            )
                            if page_input:
                                try:
                                    # Parse page numbers
                                    pages = []
                                    for part in page_input.split(','):
                                        part = part.strip()
                                        if '-' in part:
                                            start, end = map(int, part.split('-'))
                                            pages.extend(range(start, end + 1))
                                        else:
                                            pages.append(int(part))
                                    st.session_state.selected_pages[uploaded_file.name] = sorted(set(pages))
                                    st.success(f"Selected pages: {st.session_state.selected_pages[uploaded_file.name]}")
                                except:
                                    st.error("Invalid page format. Use comma-separated numbers or ranges (e.g., 1,2,3 or 1-5)")
                # Step 3: Process ALL selected pages from ALL PDFs
                st.divider()
                st.markdown("### 🚀 Step 3: Process All Selected Pages")
                # Calculate total selected pages across all files
                total_selected = 0
                files_with_selections = []
                for fname in [f.name for f in uploaded_files]:
                    selected = st.session_state.selected_pages.get(fname, [])
                    if selected:
                        total_selected += len(selected)
                        files_with_selections.append(f"{fname}: {len(selected)} pages")
                if total_selected > 0:
                    st.info(f"📋 Ready to process {total_selected} total page(s) from {len(files_with_selections)} file(s)")
                    # Show breakdown
                    with st.expander("📊 Selection Breakdown"):
                        for item in files_with_selections:
                            st.write(f"• {item}")
                if st.button("🚀 Extract Tables from All Selected Pages", type="primary", use_container_width=True, key="extract_all_tables_btn"):
                    # Get AWS clients
                    aws_session = st.session_state.aws_session
                    textract = aws_session.client('textract')
                    s3 = aws_session.client(
                        's3',
                        endpoint_url=f'https://s3.{cred_data["AWS_REGION"]}.amazonaws.com'
                    )
                    # Process each file
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    files_to_process = [f for f in uploaded_files if st.session_state.selected_pages.get(f.name, [])]
                    for idx, file in enumerate(files_to_process):
                        selected_pages_list = st.session_state.selected_pages.get(file.name, [])
                        status_text.text(f"Processing {idx+1}/{len(files_to_process)}: {file.name} ({len(selected_pages_list)} pages)")
                        try:
                            # Reset file pointer
                            file.seek(0)
                            # Process selected pages
                            all_blocks = process_single_file(
                                file,
                                textract,
                                s3,
                                cred_data,
                                selected_pages=selected_pages_list
                            )
                            # Extract tables
                            tables = extract_tables_from_blocks(all_blocks)
                            # Store results
                            st.session_state.processed_files[file.name] = {
                                'Blocks': all_blocks,
                                'tables': tables,
                                'timestamp': datetime.now(),
                                'selected_pages': selected_pages_list
                            }
                        except Exception as e:
                            st.error(f"Error processing {file.name}: {str(e)}")
                        progress_bar.progress((idx + 1) / len(files_to_process))
                    status_text.text("✅ All files processed!")
                    st.success(f"✅ Successfully processed {len(files_to_process)} file(s)!")
                    # Clear PDF previews to save memory
                    for fname in [f.name for f in uploaded_files]:
                        if fname in st.session_state.pdf_pages:
                            del st.session_state.pdf_pages[fname]
                else:
                    st.warning("⚠️ Please select at least one page from any PDF to process")
                # Display results
                if st.session_state.processed_files:
                    st.divider()
                    st.markdown('<div class="section-header">📊 Extracted Tables</div>', unsafe_allow_html=True)
                    for file_name, file_data in st.session_state.processed_files.items():
                        with st.expander(f"📄 {file_name} ({len(file_data['tables'])} tables from pages {file_data.get('selected_pages', 'all')})", expanded=True):
                            if file_data['tables']:
                                for idx, table in enumerate(file_data['tables']):
                                    st.markdown(f"Table {idx + 1}")
                                    df = pd.DataFrame(table)
                                    st.dataframe(df, use_container_width=True)
                            else:
                                st.info("No tables found in selected pages")
                # Combine JSONs feature
                st.divider()
                st.markdown("### 🔗 Combine Multiple JSONs")
                if len(st.session_state.processed_files) > 1:
                    st.markdown("Select files to combine into one table:")
                    # Let user select which files to combine
                    files_to_combine = st.multiselect(
                        "Choose files:",
                        options=list(st.session_state.processed_files.keys()),
                        default=list(st.session_state.processed_files.keys()),
                        key="combine_files_select"
                    )
                    if len(files_to_combine) > 1:
                        combined_name = st.text_input(
                            "Name for combined file:",
                            value="Combined_Cash_Flow",
                            key="combined_name_input"
                        )
                        if st.button("🔗 Combine Selected Files", type="primary", key="combine_btn"):
                            # Combine all tables from selected files
                            all_combined_tables = []
                            all_blocks = []
                            for file_name in files_to_combine:
                                file_data = st.session_state.processed_files[file_name]
                                all_combined_tables.extend(file_data['tables'])
                                all_blocks.extend(file_data['Blocks'])
                            # Create combined JSON
                            combined_json_name = f"{combined_name}.pdf"
                            st.session_state.processed_files[combined_json_name] = {
                                'tables': all_combined_tables,
                                'Blocks': all_blocks,
                                'timestamp': datetime.now(),
                                'selected_pages': 'combined',
                                'combined_from': files_to_combine
                            }
                            st.success(f"✅ Combined {len(files_to_combine)} files into '{combined_json_name}'!")
                            st.info(f"📊 Total tables: {len(all_combined_tables)}")
                            st.rerun()
                    else:
                        st.info("Select at least 2 files to combine")
                else:
                    st.info("Need at least 2 processed files to combine")
                # Export options
                st.divider()
                st.markdown("### 📥 Export Options")
                # Textract JSON format (compatibility mode)
                compatibility_mode = st.checkbox(
                    "🔄 Compatibility Mode (Textract JSON format)",
                    value=True,
                    help="Export in format compatible with Cash Flow Workflow: {'Blocks': [...]}"
                )
                # Generate and download JSON
                st.write("Download Textract JSON for Cash Flow Workflow:")
                for file_name, file_data in st.session_state.processed_files.items():
                    if compatibility_mode:
                        # Simple format for compatibility
                        json_data = {'Blocks': file_data['Blocks']}
                    else:
                        # Detailed format with metadata
                        json_data = {
                            "document_info": {
                                "name": file_name,
                                "processing_date": file_data['timestamp'].isoformat(),
                                "tables_count": len(file_data['tables']),
                                "selected_pages": file_data.get('selected_pages', 'all')
                            },
                            "Blocks": file_data['Blocks']
                        }
                    # Download button
                    json_str = json.dumps(json_data, indent=2)
                    st.download_button(
                        label=f"📥 Download {file_name}.json",
                        data=json_str,
                        file_name=f"{file_name.split('.')[0]}textract.json",
                        mime="application/json",
                        key=f"json{file_name}",
                        use_container_width=True
                    )
                # Navigation to Cash Flow Workflow
                st.divider()
                st.markdown("### ➡️ Next Steps")
                col1, col2, col3 = st.columns(3)
                with col1:
                    if st.button("🚀 Push All JSONs to Cash Flow Workflow", type="primary", use_container_width=True, key="push_jsons_btn"):
                        # Convert processed files to JSON format and store in cash flow workflow
                        for file_name, file_data in st.session_state.processed_files.items():
                            json_data = {'Blocks': file_data['Blocks']}
                            st.session_state.cfs_json_files[file_name] = {
                                'data': json_data,
                                'tables': file_data['tables'],
                                'source': 'pushed',
                                'timestamp': file_data['timestamp']
                            }
                        st.session_state.app_mode = 'main'
                        st.session_state.workflow_step = "📋 Cash Flow Prep & AI Classification"
                        st.success(f"✅ Pushed {len(st.session_state.processed_files)} JSON(s) to Cash Flow Workflow!")
                        time.sleep(1)
                        st.rerun()
                with col2:
                    if st.button("💰 Go to Cash Flow Workflow", use_container_width=True):
                        st.session_state.app_mode = 'main'
                        st.session_state.workflow_step = "📋 Cash Flow Prep & AI Classification"
                        st.rerun()
                with col3:
                    if st.button("🗑️ Clear Results", use_container_width=True):
                        st.session_state.processed_files = {}
                        st.session_state.pdf_pages = {}
                        st.session_state.selected_pages = {}
                        st.rerun()
            else:
                st.info("👆 Click 'Test AWS Connection' to start")
        except Exception as e:
            st.error(f"Setup error: {str(e)}")
    elif st.session_state.app_mode == 'ciq':
        # CIQ Template Upload Mode
        st.markdown('<div class="main-header">📤 CIQ Template Upload Center<br>Populate CIQ templates with your mapped financial data</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-header">Template Processing</div>', unsafe_allow_html=True)
        st.info("""
This section allows you to populate CIQ templates with your mapped financial data.
Requirements:
* CIQ Template file (Annual or Quarterly)
* Completed financial statement data files with:
  * 'Standardized - [Statement]' sheet containing CIQ column
  * 'As Presented - [Statement]' sheet
""")
        # Create tabs for Annual and Quarterly
        template_tab1, template_tab2 = st.tabs(["📅 Annual Upload Template", "📊 Quarterly Upload Template"])
        with template_tab1:
            st.markdown('<div class="section-header">Annual Template Processing</div>', unsafe_allow_html=True)
            process_ciq_template("Annual", session)
        with template_tab2:
            st.markdown('<div class="section-header">Quarterly Template Processing</div>', unsafe_allow_html=True)
            process_ciq_template("Quarterly", session)
        st.divider()
        st.markdown("### 📝 Important Notes")
        st.markdown("""
* The CIQ column is automatically included when you generate output in the 'Map to Standards' tab
* Sheet names must match exactly: 'Standardized - Cash Flow' and 'As Presented - Cash Flow'
* The template will be populated based on CIQ ID matching
* Original formulas in the template are preserved where possible
* Check that your date columns match between the template and your data files
""")
    elif st.session_state.app_mode == 'main':
        st.markdown('<div class="main-header">💰 Cash Flow Statement Processing Workflow<br> Classify, aggregate, map, and manage dictionary</div>', unsafe_allow_html=True)
        # Workflow navigation dropdown
        workflow_options = [
            "📋 Cash Flow Prep & AI Classification",
            "📊 Aggregate Data",
            "🔗 Map to Standards",
            "📚 Data Dictionary Management"
        ]
        current_step = st.session_state.workflow_step
        selected_step = st.selectbox(
            "Navigate Workflow",
            options=workflow_options,
            index=workflow_options.index(current_step),
            label_visibility="collapsed"
        )
        if selected_step != current_step:
            st.session_state.workflow_step = selected_step
            st.rerun()
        st.divider()
        if st.session_state.workflow_step == "📋 Cash Flow Prep & AI Classification":
            # Original's json load and per file processing
            st.markdown('<div class="section-header">Cash Flow Prep & AI Classification</div>', unsafe_allow_html=True)
            # Two input options
            st.markdown("### 📂 Load JSON Files")
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("#### Option 1: Upload Manually")
                uploaded_jsons = st.file_uploader(
                    "Upload Textract JSON file(s):",
                    type=['json'],
                    accept_multiple_files=True,
                    key="json_uploader"
                )
                if uploaded_jsons:
                    for uploaded_json in uploaded_jsons:
                        if uploaded_json.name not in st.session_state.cfs_json_files:
                            try:
                                json_data = json.load(uploaded_json)
                                if 'Blocks' in json_data:
                                    tables = extract_tables_from_textract(json_data)
                                    st.session_state.cfs_json_files[uploaded_json.name] = {
                                        'data': json_data,
                                        'tables': tables,
                                        'source': 'uploaded',
                                        'timestamp': datetime.now()
                                    }
                                    st.success(f"✅ Loaded: {uploaded_json.name}")
                            except Exception as e:
                                st.error(f"Error loading {uploaded_json.name}: {str(e)}")
            with col2:
                st.markdown("#### Option 2: Pushed from Document Processing")
                pushed_count = len([f for f in st.session_state.cfs_json_files.values() if f.get('source') == 'pushed'])
                if pushed_count > 0:
                    st.info(f"📋 {pushed_count} JSON file(s) pushed from Document Processing")
                else:
                    st.info("No files pushed yet. Process PDFs in Document Processing tab and click 'Push All JSONs'.")
            # Display all JSONs
            if st.session_state.cfs_json_files:
                st.divider()
                st.markdown(f"### 📊 Processing {len(st.session_state.cfs_json_files)} JSON File(s)")
                # Get AI model for classification
                model = MODELS[0] # Default model
                session = get_active_session()
                # Process each JSON
                for json_idx, (file_name, file_info) in enumerate(st.session_state.cfs_json_files.items()):
                    if file_name not in st.session_state.cfs_prep_expander_state:
                        st.session_state.cfs_prep_expander_state[file_name] = True
                    with st.expander(f"📄 {file_name} ({len(file_info['tables'])} table(s))", expanded=st.session_state.cfs_prep_expander_state[file_name]):
                        st.markdown(f"Source: {file_info['source'].title()} | Loaded: {file_info['timestamp'].strftime('%Y-%m-%d %H:%M')}")
                        # Step 1: Select Table(s) - Support multiple table selection
                        st.markdown("#### 📋 Step 1: Select Table(s)")
                        # Initialize selected tables for this file
                        if file_name not in st.session_state.cfs_selected_tables:
                            st.session_state.cfs_selected_tables[file_name] = [0] if len(file_info['tables']) > 0 else []
                        if len(file_info['tables']) > 1:
                            st.markdown("Choose which table(s) to process:")
                            st.info("💡 Select multiple tables to combine them into one table for processing")
                            # Show checkboxes for each table
                            selected_indices = []
                            for idx, table in enumerate(file_info['tables']):
                                df = pd.DataFrame(table)
                                col1, col2 = st.columns([1, 4])
                                with col1:
                                    is_checked = idx in st.session_state.cfs_selected_tables[file_name]
                                    if st.checkbox(
                                        f"Include",
                                        value=is_checked,
                                        key=f"table_check_{json_idx}_{idx}"
                                    ):
                                        selected_indices.append(idx)
                                with col2:
                                    st.markdown(f"Table {idx + 1} ({df.shape[0]} rows x {df.shape[1]} cols)")
                                    with st.expander("👁️ Preview", expanded=False):
                                        st.dataframe(df.head(10), use_container_width=True)
                            # Update stored selection
                            if selected_indices:
                                st.session_state.cfs_selected_tables[file_name] = selected_indices
                            else:
                                st.warning("⚠️ Please select at least one table")
                                st.session_state.cfs_selected_tables[file_name] = [0]
                                selected_indices = [0]
                            # Show what will be processed
                            if len(selected_indices) > 1:
                                st.success(f"✅ {len(selected_indices)} tables will be combined vertically")
                            else:
                                st.info(f"📋 Processing Table {selected_indices[0] + 1}")
                        else:
                            # Single table - no selection needed
                            st.session_state.cfs_selected_tables[file_name] = [0]
                            selected_indices = [0]
                            preview_df = pd.DataFrame(file_info['tables'][0])
                            st.markdown(f"Table Preview ({preview_df.shape[0]} rows x {preview_df.shape[1]} cols)")
                            st.dataframe(preview_df.head(10), use_container_width=True)
                        # Get the combined DataFrame from selected tables
                        selected_table_indices = st.session_state.cfs_selected_tables[file_name]
                        if len(selected_table_indices) > 1:
                            # Combine multiple tables vertically
                            combined_dfs = []
                            for idx in selected_table_indices:
                                table_data = file_info['tables'][idx]
                                combined_dfs.append(pd.DataFrame(table_data))
                            # Concatenate vertically, reset index
                            df = pd.concat(combined_dfs, ignore_index=True)
                            # Show combined preview
                            with st.expander("📊 Preview of Combined Table", expanded=True):
                                st.markdown(f"Combined: {df.shape[0]} rows x {df.shape[1]} cols (from {len(selected_table_indices)} tables)")
                                st.dataframe(df.head(15), use_container_width=True)
                        else:
                            # Single table selected
                            table_data = file_info['tables'][selected_table_indices[0]]
                            df = pd.DataFrame(table_data)
                        st.divider()
                        # ===== STEP 1: Select Account Column =====
                        st.markdown("#### 📌 Step 1: Select Account Column")
                        # Initialize account column selection for this file
                        if file_name not in st.session_state.cfs_account_column:
                            # Auto-detect first text column as default
                            default_col_idx = 0
                            for col_idx, col in enumerate(df.columns):
                                if df[col].dtype == 'object' and not df[col].str.contains(r'^\d+$', na=False).all():
                                    default_col_idx = col_idx
                                    break
                            st.session_state.cfs_account_column[file_name] = default_col_idx
                        # Let user select account column
                        account_col_idx = st.selectbox(
                            "Select which column contains account names:",
                            options=range(len(df.columns)),
                            format_func=lambda x: f"Column {x}: '{df.columns[x]}'",
                            index=st.session_state.cfs_account_column[file_name],
                            key=f"account_col_select_{json_idx}"
                        )
                        st.session_state.cfs_account_column[file_name] = account_col_idx
                        account_column = df.columns[account_col_idx]
                        st.success(f"✅ Using column '{account_column}' as account names")
                        # Show preview of account column
                        with st.expander("👁️ Preview Account Column"):
                            st.dataframe(
                                df[[account_column]].head(10),
                                use_container_width=True
                            )
                        st.divider()
                        # ===== STEP 2: Select Value Columns =====
                        st.markdown("#### 📊 Step 2: Select Value Columns")
                        all_non_account_cols = [c for c in df.columns if c != account_column]
                        if file_name not in st.session_state.cfs_selected_value_cols:
                            st.session_state.cfs_selected_value_cols[file_name] = all_non_account_cols
                        selected_value_cols = st.multiselect(
                            "Select columns to include as values:",
                            options=all_non_account_cols,
                            default=st.session_state.cfs_selected_value_cols[file_name],
                            key=f"value_cols_{json_idx}"
                        )
                        if selected_value_cols:
                            st.session_state.cfs_selected_value_cols[file_name] = selected_value_cols
                        else:
                            st.warning("⚠️ Please select at least one value column")
                            st.session_state.cfs_selected_value_cols[file_name] = all_non_account_cols
                            selected_value_cols = all_non_account_cols
                        st.divider()
                        # ===== STEP 3: Classify Accounts =====
                        st.markdown("#### 🏷️ Step 3: Classify Accounts")
                        
                        # Get the selected value columns from Step 2
                        selected_value_cols = st.session_state.cfs_selected_value_cols.get(file_name, [])
                        
                        if st.button(f"🤖 Classify with AI", key=f"classify_{json_idx}"):
                            with st.spinner("Classifying accounts with AI..."):
                                classifications = classify_cash_flow_with_ai(df, account_column, model, session)
                                st.session_state.cfs_classifications[file_name] = classifications
                            st.success(f"✅ Classified {len(classifications)} accounts!")
                        
                        # Show classifications if available
                        if file_name in st.session_state.cfs_classifications:
                            classifications = st.session_state.cfs_classifications[file_name]
                            
                            # Create classification display DataFrame
                            class_df = df.copy()
                            class_df['Label'] = class_df.index.map(
                                lambda idx: classifications.get(idx, {}).get('category', 'Unclassified')
                            )
                            class_df['Confidence'] = class_df.index.map(
                                lambda idx: f"{classifications.get(idx, {}).get('confidence', 0):.0%}"
                            )
                            
                            # Initialize temp classifications if not exists (to prevent jitter)
                            temp_class_key = f"temp_classifications_{file_name}"
                            if temp_class_key not in st.session_state:
                                st.session_state[temp_class_key] = classifications.copy()
                            
                            # Build columns to display: only account column and SELECTED value columns
                            display_cols = [account_column] + selected_value_cols + ['Label', 'Confidence']
                            
                            # Prepare column config to make only Label editable
                            column_config = {
                                'Label': st.column_config.SelectboxColumn(
                                    'Label',
                                    options=['Cash from Operations', 'Cash from Investing', 'Cash from Financing', 'Cash from other', 'Subtotal', 'Skip', ''],
                                    required=False
                                ),
                                'Confidence': st.column_config.TextColumn(disabled=True)
                            }
                            
                            # Make all other columns read-only
                            for col in [account_column] + selected_value_cols:
                                if class_df[col].dtype == 'object':
                                    column_config[col] = st.column_config.TextColumn(disabled=True)
                                else:
                                    column_config[col] = st.column_config.NumberColumn(disabled=True)
                            
                            st.markdown("**📊 Classification Results** (Edit labels as needed)")
                            st.info(f"💡 Showing {len(selected_value_cols)} selected value column(s). Adjust in Step 2 if needed.")
                            
                            # Use a form to batch updates and prevent jittering
                            with st.form(key=f"classification_form_{json_idx}"):
                                edited_class_df = st.data_editor(
                                    class_df[display_cols],
                                    column_config=column_config,
                                    hide_index=True,
                                    use_container_width=True,
                                    key=f"class_editor_{json_idx}"
                                )
                                
                                # Submit button to save changes
                                col1, col2 = st.columns([1, 3])
                                with col1:
                                    save_changes = st.form_submit_button("💾 Save Changes", type="primary", use_container_width=True)
                                with col2:
                                    st.markdown("*Changes will be saved when you click the button*")
                            
                            # Only update classifications when form is submitted
                            if save_changes:
                                # Update classifications based on edits
                                changes_made = False
                                for idx, row in edited_class_df.iterrows():
                                    if idx in classifications:
                                        old_label = classifications[idx]['category']
                                        new_label = row['Label']
                                        if old_label != new_label:
                                            classifications[idx]['category'] = new_label
                                            changes_made = True
                                
                                if changes_made:
                                    st.session_state.cfs_classifications[file_name] = classifications
                                    st.success("✅ Classifications updated successfully!")
                                    st.rerun()
                                else:
                                    st.info("ℹ️ No changes detected")
                        
                        # Update numeric_cols for later use
                        numeric_cols = selected_value_cols
                        st.divider()
                        # ===== STEP 4: Name Columns with Dropdowns and Manual Input =====
                        st.markdown("#### 📅 Step 4: Name Columns")
                        # Bug 2 Fix: Initialize column names only for selected columns and clean up deselected ones
                        if file_name not in st.session_state.cfs_column_names:
                            st.session_state.cfs_column_names[file_name] = {
                                col: f"Period_{idx}" for idx, col in enumerate(selected_value_cols)
                            }
                        else:
                            # Clean up column names for deselected columns
                            current_col_names = st.session_state.cfs_column_names[file_name]
                            # Remove columns that are no longer selected
                            st.session_state.cfs_column_names[file_name] = {
                                col: name for col, name in current_col_names.items()
                                if col in selected_value_cols
                            }
                            # Add default names for newly selected columns
                            for idx, col in enumerate(selected_value_cols):
                                if col not in st.session_state.cfs_column_names[file_name]:
                                    st.session_state.cfs_column_names[file_name][col] = f"Period_{idx}"
                        st.markdown("Define what each numeric column represents:")
                        st.markdown("Choose from predefined periods or enter a custom name")
                        # Create period options
                        fiscal_year_options = [f"FY{year}" for year in range(2018, 2027)]
                        quarterly_options = [f"FY{year}_Q{quarter}" for year in range(2018, 2027) for quarter in range(1, 5)]
                        ytd_options = [f"YTD{quarter}{year}" for year in range(2018, 2027) for quarter in range(1, 4)]
                        dropdown_options = [''] + ['Custom'] + fiscal_year_options + quarterly_options + ytd_options
                        # Display column naming interface
                        numeric_cols = selected_value_cols
                        for idx, col in enumerate(numeric_cols):
                            col1, col2 = st.columns([1, 1])
                            with col1:
                                # Dropdown for predefined periods
                                selected_dropdown = st.selectbox(
                                    f"Period preset for '{col}':",
                                    dropdown_options,
                                    key=f"dropdown_{json_idx}_{col}",
                                    index=0
                                )
                            with col2:
                                # Text input for manual/custom entry
                                current_value = st.session_state.cfs_column_names[file_name].get(col, f"Period_{idx}")
                                # If dropdown is selected and not empty or 'Custom', use dropdown value
                                if selected_dropdown and selected_dropdown != 'Custom':
                                    display_value = selected_dropdown
                                    is_disabled = True
                                else:
                                    display_value = current_value
                                    is_disabled = False
                                manual_input = st.text_input(
                                    f"Or enter custom name:",
                                    value=display_value,
                                    key=f"manual_{json_idx}_{col}",
                                    disabled=is_disabled,
                                    help="Select 'Custom' in dropdown to enable manual entry"
                                )
                                # Store the final column name
                                if selected_dropdown and selected_dropdown != 'Custom':
                                    st.session_state.cfs_column_names[file_name][col] = selected_dropdown
                                else:
                                    st.session_state.cfs_column_names[file_name][col] = manual_input
                        st.divider()
                        # ===== STEP 4.5: Units Conversion =====
                        st.markdown("#### 🔢 Step 4.5: Units Conversion")
                        # Initialize units conversion for this file
                        if file_name not in st.session_state.cfs_units_conversion:
                            st.session_state.cfs_units_conversion[file_name] = "Actuals"
                        st.markdown("Select the units for your financial values:")
                        col1, col2 = st.columns([2, 3])
                        with col1:
                            selected_units = st.radio(
                                "Conversion units:",
                                ["Actuals", "Thousands", "Millions", "Billions"],
                                index=["Actuals", "Thousands", "Millions", "Billions"].index(
                                    st.session_state.cfs_units_conversion[file_name]
                                ),
                                key=f"units_{json_idx}",
                                help="Values will be converted to the selected units"
                            )
                            st.session_state.cfs_units_conversion[file_name] = selected_units
                        with col2:
                            # Show example
                            example_value = 1234567.89
                            conversion_factors_display = {
                                "Actuals": (1, "1,234,567.89"),
                                "Thousands": (1000, "1,234.57"),
                                "Millions": (1000000, "1.23"),
                                "Billions": (1000000000, "0.00")
                            }
                            factor, display = conversion_factors_display[selected_units]
                            st.info(f"💡 Example: $1,234,567.89 → ${display} ({selected_units})")
                        st.divider()
                        # ===== STEP 5: Results & Actions =====
                        st.markdown("#### ✅ Step 5: Results & Actions")
                        if file_name in st.session_state.cfs_classifications and st.session_state.cfs_column_names.get(file_name):
                            # Bug 2 Fix: Get the user's selected value columns from Step 2
                            selected_value_cols = st.session_state.cfs_selected_value_cols.get(file_name, [])

                            # Create final processed DataFrame with ONLY account column and selected value columns
                            final_df = df[[account_column] + selected_value_cols].copy()
                            final_df['Label'] = final_df.index.map(
                                lambda idx: st.session_state.cfs_classifications[file_name].get(idx, {}).get('category', 'Unclassified')
                            )

                            # Rename columns - only rename the account column and selected value columns
                            col_rename = {account_column: 'Account'}
                            # Only include renames for columns that are actually in selected_value_cols
                            for orig_col in selected_value_cols:
                                if orig_col in st.session_state.cfs_column_names.get(file_name, {}):
                                    col_rename[orig_col] = st.session_state.cfs_column_names[file_name][orig_col]

                            final_df = final_df.rename(columns=col_rename)

                            # Reorder columns: Label, Account, then ONLY the renamed selected periods
                            period_cols = [col_rename.get(c, c) for c in selected_value_cols]
                            cols_order = ['Label', 'Account'] + period_cols
                            final_df = final_df[cols_order]
                            # Filter to only keep valid labels (remove blank, unclassified, skip, etc.)
                            valid_labels = ['Cash from Operations', 'Cash from Investing', 'Cash from Financing', 'Cash from other', 'Subtotal']
                            final_df = final_df[final_df['Label'].isin(valid_labels)]
                            # Apply units conversion
                            conversion_factors = {
                                "Actuals": 1,
                                "Thousands": 1000,
                                "Millions": 1000000,
                                "Billions": 1000000000
                            }
                            conversion_factor = conversion_factors[st.session_state.cfs_units_conversion[file_name]]
                            # Convert numeric columns
                            numeric_cols = [c for c in final_df.columns if c not in ['Label', 'Account']]
                            for col in numeric_cols:
                                final_df[col] = final_df[col].apply(clean_numeric_value)
                                if conversion_factor != 1:
                                    final_df[col] = final_df[col] * conversion_factor
                            # Sort by label order
                            final_df = sort_by_label_and_account(final_df, 'Account')
                            # Store in session state
                            st.session_state.cfs_processed_data[file_name] = final_df
                            # Display results
                            st.markdown("📊 Processed Data Preview:")
                            st.info(f"📏 Units: {st.session_state.cfs_units_conversion[file_name]} | 📋 Rows: {len(final_df)} (only valid labels)")
                            st.dataframe(final_df, use_container_width=True)
                            # Action buttons
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                # Download as Excel (only valid labels)
                                buffer = io.BytesIO()
                                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                    final_df.to_excel(writer, index=False, sheet_name='Cash Flow')
                                st.download_button(
                                    label="📥 Download Excel",
                                    data=buffer.getvalue(),
                                    file_name=f"{file_name.split('.')[0]}processed{st.session_state.cfs_units_conversion[file_name]}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"download_excel_{json_idx}",
                                    use_container_width=True
                                )
                            with col2:
                                # Display full results
                                if st.button("👁️ View Full Results", key=f"view_results_{json_idx}", use_container_width=True):
                                    st.session_state[f"show_full_{file_name}"] = not st.session_state.get(f"show_full_{file_name}", False)
                            with col3:
                                # Mark/Unmark as ready for aggregation
                                is_ready = file_name in st.session_state.cfs_ready_for_aggregation
                                button_label = "✅ Marked Ready" if is_ready else "➡️ Mark Ready"
                                button_type = "secondary" if is_ready else "primary"
                                if st.button(button_label, key=f"next_step_{json_idx}", type=button_type, use_container_width=True):
                                    if is_ready:
                                        st.session_state.cfs_ready_for_aggregation.discard(file_name)
                                        st.success("❌ Unmarked from aggregation step!")
                                    else:
                                        st.session_state.cfs_ready_for_aggregation.add(file_name)
                                        st.success("✅ Marked as ready for aggregation step!")
                                    st.rerun()
                            # Show full results if toggled
                            if st.session_state.get(f"show_full_{file_name}", False):
                                st.markdown("Complete Processed Data:")
                                st.dataframe(final_df, use_container_width=True, height=600)
                        else:
                            st.warning("⚠️ Please complete classification and column naming above")
                # Summary section
                st.divider()
                st.markdown("### 📊 Summary")
                processed_count = len(st.session_state.cfs_processed_data.keys())
                ready_count = len(st.session_state.cfs_ready_for_aggregation)
                total_count = len(st.session_state.cfs_json_files)
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Files", total_count)
                with col2:
                    st.metric("Processed", processed_count)
                with col3:
                    st.metric("Ready for Aggregation", ready_count, delta=f"{ready_count - processed_count}" if ready_count != processed_count else None)
                with col4:
                    st.metric("Pending", total_count - processed_count)
                # Show list of ready files with ability to unmark
                if ready_count > 0:
                    st.divider()
                    st.markdown("#### ✅ Files Marked Ready for Aggregation")
                    for idx, ready_file in enumerate(sorted(st.session_state.cfs_ready_for_aggregation)):
                        col1, col2 = st.columns([4, 1])
                        with col1:
                            st.markdown(f"{idx + 1}. {ready_file}")
                        with col2:
                            if st.button("❌ Unmark", key=f"unmark_{ready_file}_{idx}", use_container_width=True):
                                st.session_state.cfs_ready_for_aggregation.discard(ready_file)
                                st.rerun()
                # Continue to aggregation button
                if ready_count > 0:
                    st.divider()
                    st.success(f"🎉 {ready_count} file(s) ready for aggregation!")
                    if st.button("🚀 Proceed to Aggregation Step", type="primary", use_container_width=True):
                        st.session_state.workflow_step = "📊 Aggregate Data"
                        st.rerun()
            else:
                st.info("📂 No JSON files loaded. Upload files or push them from Document Processing.")
            # Clear all button
            st.divider()
            if st.button("🗑️ Clear All and Start Over", use_container_width=True):
                st.session_state.cfs_json_files = {}
                st.session_state.cfs_classifications = {}
                st.session_state.cfs_column_names = {}
                st.session_state.cfs_selected_tables = {}
                st.session_state.cfs_processed_data = {}
                st.session_state.cfs_account_column = {}
                st.session_state.cfs_units_conversion = {}
                st.session_state.cfs_ready_for_aggregation = set()
                st.session_state.cfs_selected_value_cols = {}
                st.session_state.cfs_prep_expander_state = {}
                st.rerun()
        elif st.session_state.workflow_step == "📊 Aggregate Data":
            st.markdown('<div class="section-header">Aggregate Data</div>', unsafe_allow_html=True)
            if st.session_state.show_removed_rows_message and st.session_state.removed_zero_rows:
                st.success(f"✅ Successfully removed {len(st.session_state.removed_zero_rows)} zero-value rows from the dataset")
                with st.expander("📋 Click to view details of removed rows", expanded=False):
                    removed_df = pd.DataFrame(st.session_state.removed_zero_rows)
                    st.dataframe(removed_df, use_container_width=True)
                if st.button("✓ Dismiss Message", key="dismiss_removed_rows"):
                    st.session_state.show_removed_rows_message = False
                    st.session_state.removed_zero_rows = []
                    st.rerun()
            ready_count = len(st.session_state.cfs_ready_for_aggregation)
            # Bug 1 Fix: Only show aggregate button when data hasn't been aggregated yet
            if ready_count > 0 and st.session_state.cfs_aggregated_data is None:
                if st.button("🚀 Aggregate Ready Files", type="primary", use_container_width=True):
                    ready_dfs = []
                    for fname in st.session_state.cfs_ready_for_aggregation:
                        if fname in st.session_state.cfs_processed_data:
                            ready_dfs.append(st.session_state.cfs_processed_data[fname])
                    if ready_dfs:
                        with st.spinner("Aggregating data..."):
                            combined_df = pd.concat(ready_dfs, ignore_index=True)
                            aggregated_df = aggregate_data(combined_df)
                            st.session_state.cfs_aggregated_data = sort_by_label_and_account(aggregated_df, 'Account')
                        st.rerun()
                    else:
                        st.warning("No processed data found for ready files")
            elif ready_count == 0 and st.session_state.cfs_aggregated_data is None:
                st.info("No files ready for aggregation. Mark files ready in the Prep step.")
            if st.session_state.cfs_aggregated_data is not None:
                st.markdown('<div class="main-header">📊 Aggregated Results</div>', unsafe_allow_html=True)
                st.write("### Aggregated Data:")
                st.dataframe(st.session_state.cfs_aggregated_data, use_container_width=True)
                aggregated_table = st.session_state.cfs_aggregated_data
                zero_rows = check_all_zeroes(aggregated_table)
                zero_count = zero_rows.sum()
                if zero_count > 0:
                    st.warning(f"⚠️ Found {zero_count} rows where ALL numeric columns are zero.")
                    with st.expander(f"Preview {zero_count} zero-value rows"):
                        zero_df = aggregated_table[zero_rows]
                        st.dataframe(zero_df)
                    if st.button("🗑️ Remove Zero-Value Rows", type="secondary", key="remove_zero_rows_btn"):
                        st.session_state.removed_zero_rows = zero_df[['Label', 'Account']].to_dict('records')
                        st.session_state.show_removed_rows_message = True
                        cleaned_table = aggregated_table[~zero_rows].reset_index(drop=True)
                        st.session_state.cfs_aggregated_data = cleaned_table
                        st.rerun()
                else:
                    st.success("✅ No rows with all zero values found in the current dataset.")
                st.divider()
                col1, col2 = st.columns(2)
                with col1:
                    excel_file = io.BytesIO()
                    st.session_state.cfs_aggregated_data.to_excel(excel_file, index=False)
                    excel_file.seek(0)
                    st.download_button(
                        "📥 Download Aggregated Data",
                        excel_file,
                        "aggregated_cash_flow.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                with col2:
                    if st.button("➡️ Proceed to Mapping", type="primary", key="proceed_to_mapping_btn"):
                        st.session_state.workflow_step = "🔗 Map to Standards"
                        st.rerun()
            else:
                st.info("No aggregated data yet. Mark files ready in Prep step and aggregate.")
        elif st.session_state.workflow_step == "🔗 Map to Standards":
            st.markdown('<div class="section-header">Map to Standards</div>', unsafe_allow_html=True)
            if st.session_state.cfs_aggregated_data is None:
                st.info("Please aggregate data in the previous step first.")
                return
            # The mapping code from the new, with optimizations
            if st.session_state.show_success_message and st.session_state.just_added_mappings:
                st.success(f"✅ Successfully added {len(st.session_state.just_added_mappings)} new mapping(s) to dictionary!")
                with st.expander("View recently added mappings", expanded=True):
                    added_df = pd.DataFrame(st.session_state.just_added_mappings)
                    st.dataframe(added_df, use_container_width=True)
                st.info("These accounts will now match automatically in future mappings!")
                if st.button("Clear message", key="clear_mapping_message"):
                    st.session_state.show_success_message = False
                    st.session_state.just_added_mappings = []
                    st.rerun()
            st.markdown("#### 📋 Workflow")
            workflow_cols = st.columns(4)
            with workflow_cols[0]:
                st.info("1️⃣ Run Mapping")
            with workflow_cols[1]:
                st.info("2️⃣ Review Matches")
            with workflow_cols[2]:
                st.info("3️⃣ Generate Output")
            with workflow_cols[3]:
                st.success("4️⃣ Update Dictionary")
            if st.session_state.cfs_dictionary_data is None:
                dict_data = load_dictionary_cached(cache_key)
                if dict_data is not None:
                    st.session_state.cfs_dictionary_data = dict_data
            aggregated = st.session_state.cfs_aggregated_data
            col1, col2, col3 = st.columns(3)
            with col1:
                st.session_state.company_name = st.text_input("Company Name", value=st.session_state.company_name)
            with col2:
                currency = st.selectbox("Currency", ["U.S. Dollar", "Euro", "British Pound Sterling", "Japanese Yen"])
            with col3:
                magnitude = st.selectbox("Magnitude", ["Actuals", "Thousands", "Millions", "Billions"])
            st.markdown("### Mapping Options")
            col1, col2 = st.columns(2)
            with col1:
                run_fuzzy = st.checkbox("Run Fuzzy Matching", value=True)
            with col2:
                run_llm = st.checkbox("Run AI Matching", value=True)
                if run_llm:
                    ai_model = st.selectbox("AI Model", MODELS)
                else:
                    ai_model = None
            st.info("💡 Tip: After generating output, you'll be able to update the data dictionary with your new mappings!")
            if st.button("🔄 Run Mapping Analysis", type="primary", use_container_width=True, key="run_mapping_analysis"):
                with st.spinner("Analyzing all matching options..."):
                    mapping_results = []
                    progress = st.progress(0)
                    for idx, row in aggregated.iterrows():
                        matches = get_all_matches(
                            row['Account'],
                            row['Label'],
                            st.session_state.cfs_dictionary_data,
                            cache_key,
                            run_llm=run_llm,
                            model=ai_model
                        )
                        mapping_results.append({
                            'index': idx,
                            'account': row['Account'],
                            'label': row['Label'],
                            'matches': matches
                        })
                        progress.progress((idx + 1) / len(aggregated))
                    st.session_state.cfs_mapping_results = mapping_results
                st.success("✅ Analysis complete! Review matches below.")
            if 'cfs_mapping_results' in st.session_state and st.session_state.cfs_mapping_results:
                st.markdown("### Review and Select Mappings")
                if 'cfs_user_selections' not in st.session_state:
                    st.session_state.cfs_user_selections = {}
                total_items = len(st.session_state.cfs_mapping_results)
                mapped_items = 0
                unmapped_items = 0
                new_items = 0
                if st.session_state.cfs_user_selections:
                    mapped_items = sum(1 for sel in st.session_state.cfs_user_selections.values() if sel['choice'] not in ['Select...', '', 'Skip/Remove', 'Leave Unmapped'])
                    unmapped_items = sum(1 for sel in st.session_state.cfs_user_selections.values() if sel['choice'] == 'Leave Unmapped')
                if st.session_state.cfs_dictionary_data is not None:
                    for result in st.session_state.cfs_mapping_results:
                        label_accounts = st.session_state.cfs_dictionary_data[
                            st.session_state.cfs_dictionary_data['LABEL'] == result['label']
                        ]['ACCOUNT'].values
                        if result['account'] not in label_accounts:
                            new_items += 1
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Items", total_items)
                with col2:
                    st.metric("Mapped", mapped_items, f"{mapped_items/total_items*100:.0f}%")
                with col3:
                    st.metric("Unmapped", unmapped_items, help="Items deliberately left unmapped")
                with col4:
                    st.metric("New Accounts", new_items, help="Accounts not in dictionary")
                st.info("For each account, choose between Fuzzy Match, AI Match, Manual Selection, or Leave Unmapped")
                st.markdown("#### Matching Settings")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    fuzzy_threshold = st.slider(
                        "Fuzzy Match Threshold",
                        min_value=50,
                        max_value=90,
                        value=70,
                        step=5,
                        help="Minimum score for fuzzy matching to be considered valid"
                    )
                with col2:
                    high_confidence_threshold = st.slider(
                        "Auto-Recommend Threshold",
                        min_value=70,
                        max_value=95,
                        value=80,
                        step=5,
                        help="Matches at or above this score are auto-recommended"
                    )
                with col3:
                    st.metric("Total Matches Found", sum(1 for r in st.session_state.cfs_mapping_results if r['matches']['fuzzy'] and r['matches']['fuzzy']['score'] >= fuzzy_threshold))
                with col4:
                    st.info("Fuzzy Matching uses text similarity to find accounts with similar names, handling variations in spelling, word order, and abbreviations")
                if 'cfs_user_selections' not in st.session_state:
                    st.session_state.cfs_user_selections = {}
                total_mapped = 0
                if st.session_state.cfs_user_selections:
                    total_mapped = sum(1 for sel in st.session_state.cfs_user_selections.values() if sel['choice'] not in ['Select...', ''])
                total_items = len(st.session_state.cfs_mapping_results)
                progress_pct = (total_mapped / total_items * 100) if total_items > 0 else 0
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.progress(progress_pct / 100, text=f"Mapping Progress: {total_mapped}/{total_items} items ({progress_pct:.0f}%)")
                with col2:
                    if progress_pct == 100:
                        st.success("✅ All items mapped!")
                    elif progress_pct > 0:
                        st.warning(f"⚠️ {total_items - total_mapped} items remaining")
                    else:
                        st.info("📝 Ready to map")
                st.markdown("### 📊 Unified Mapping Overview")
                summary_df = create_summary_df(st.session_state.cfs_mapping_results, st.session_state.cfs_dictionary_data)
                st.dataframe(
                    summary_df,
                    column_config={
                        "Fuzzy Score": st.column_config.TextColumn("Fuzzy Score", help="Text similarity score"),
                        "New": st.column_config.TextColumn("New", help="🆕 indicates accounts not in dictionary")
                    },
                    use_container_width=True,
                    height=400
                )
                with st.expander("📖 Score Legend", expanded=False):
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.success(f"High Confidence: ≥{high_confidence_threshold}%")
                    with col2:
                        st.warning(f"Medium Confidence: {fuzzy_threshold}-{high_confidence_threshold-1}%")
                    with col3:
                        st.error(f"Below Threshold: <{fuzzy_threshold}%")
                csv = summary_df.to_csv(index=False)
                st.download_button(
                    "📥 Download Mapping Summary",
                    csv,
                    "mapping_summary.csv",
                    "text/csv"
                )
                st.divider()
                show_detailed = st.checkbox("Show Detailed Selection View", value=True, help="Expand to see individual selection options for each account")
                if show_detailed:
                    st.markdown("### 🔍 Detailed Mapping Selection")
                    st.markdown("#### Quick Actions")
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        if st.button("✅ Apply All Recommendations", type="primary", key="apply_all_recommendations"):
                            for result in st.session_state.cfs_mapping_results:
                                idx = result['index']
                                matches = result['matches']
                                if matches['fuzzy'] and matches['fuzzy']['score'] >= high_confidence_threshold:
                                    choice = f"Fuzzy ({matches['fuzzy']['score']}%)"
                                elif matches['llm'] and 'error' not in matches['llm']:
                                    choice = "AI Match"
                                elif matches['fuzzy'] and matches['fuzzy']['score'] >= fuzzy_threshold:
                                    choice = f"Fuzzy ({matches['fuzzy']['score']}%)"
                                else:
                                    choice = "Select..."
                                st.session_state.cfs_user_selections[idx] = {
                                    'choice': choice,
                                    'manual': None,
                                    'fuzzy': matches['fuzzy'],
                                    'llm': matches['llm']
                                }
                            st.success("Applied recommendations to all items!")
                            st.rerun()
                    with col2:
                        if st.button("🔍 Apply All Fuzzy Matches", key="apply_all_fuzzy"):
                            for result in st.session_state.cfs_mapping_results:
                                idx = result['index']
                                matches = result['matches']
                                if matches['fuzzy']:
                                    st.session_state.cfs_user_selections[idx] = {
                                        'choice': f"Fuzzy ({matches['fuzzy']['score']}%)",
                                        'manual': None,
                                        'fuzzy': matches['fuzzy'],
                                        'llm': matches['llm']
                                    }
                            st.success("Applied fuzzy matches where available!")
                            st.rerun()
                    with col3:
                        if st.button("🤖 Apply All AI Matches", key="apply_all_ai"):
                            for result in st.session_state.cfs_mapping_results:
                                idx = result['index']
                                matches = result['matches']
                                if matches['llm'] and 'error' not in matches['llm']:
                                    st.session_state.cfs_user_selections[idx] = {
                                        'choice': "AI Match",
                                        'manual': None,
                                        'fuzzy': matches['fuzzy'],
                                        'llm': matches['llm']
                                    }
                            st.success("Applied AI matches where available!")
                            st.rerun()
                    with col4:
                        if st.button("❓ Mark Low Confidence as Unmapped", key="mark_low_confidence"):
                            for result in st.session_state.cfs_mapping_results:
                                idx = result['index']
                                matches = result['matches']
                                if not matches['fuzzy'] or matches['fuzzy']['score'] < fuzzy_threshold:
                                    st.session_state.cfs_user_selections[idx] = {
                                        'choice': "Leave Unmapped",
                                        'manual': None,
                                        'fuzzy': matches['fuzzy'],
                                        'llm': matches['llm']
                                    }
                            st.success("Marked low confidence items as unmapped!")
                            st.rerun()
                    st.divider()
                    results_by_label = {}
                    for result in st.session_state.cfs_mapping_results:
                        label = result['label']
                        if label not in results_by_label:
                            results_by_label[label] = []
                        results_by_label[label].append(result)
                    st.session_state.cfs_all_labels = list(results_by_label.keys())
                    for label in sorted(results_by_label.keys()):
                        label_results = results_by_label[label]
                        selected_count = sum(1 for r in label_results if r['index'] in st.session_state.cfs_user_selections and st.session_state.cfs_user_selections[r['index']]['choice'] not in ['Select...', ''])
                        if auto_collapse and selected_count == len(label_results):
                            expanded = False
                        elif label in st.session_state.cfs_expand_state:
                            expanded = st.session_state.cfs_expand_state[label]
                        else:
                            expanded = True
                        header_text = f"{label} ({selected_count}/{len(label_results)} mapped)"
                        with st.expander(header_text, expanded=expanded):
                            for result in label_results:
                                idx = result['index']
                                account = result['account']
                                matches = result['matches']
                                is_new_account = False
                                if st.session_state.cfs_dictionary_data is not None:
                                    label_accounts = st.session_state.cfs_dictionary_data[
                                        st.session_state.cfs_dictionary_data['LABEL'] == label
                                    ]['ACCOUNT'].values
                                    is_new_account = account not in label_accounts
                                if is_new_account:
                                    st.markdown(f"#### {account} 🆕")
                                    st.caption("This account is not in the dictionary and will be added when you update it")
                                else:
                                    st.markdown(f"#### {account}")
                                cols = st.columns([2, 2, 2, 2])
                                with cols[0]:
                                    st.markdown("🔍 Fuzzy Match")
                                    if matches['fuzzy']:
                                        fuzzy = matches['fuzzy']
                                        st.write(f"Match: {fuzzy['match']}")
                                        st.write(f"Score: {fuzzy['score']}%")
                                        st.write(f"Mnemonic: {fuzzy['mnemonic']}")
                                        if fuzzy['score'] >= high_confidence_threshold:
                                            st.success("High confidence")
                                        elif fuzzy['score'] >= fuzzy_threshold:
                                            st.warning("Medium confidence")
                                        else:
                                            st.error("Below threshold")
                                    else:
                                        st.write("No match found")
                                with cols[1]:
                                    st.markdown("🤖 AI Match")
                                    if matches['llm']:
                                        if 'error' in matches['llm']:
                                            st.error(f"Error: {matches['llm']['error']}")
                                        else:
                                            llm = matches['llm']
                                            st.write(f"Match: {llm['match']}")
                                            st.write(f"Confidence: {llm['confidence']}")
                                            st.write(f"Mnemonic: {llm['mnemonic']}")
                                    else:
                                        st.write("Not run")
                                with cols[2]:
                                    st.markdown("✏️ Manual Override")
                                    label_filtered = filter_dictionary_by_label(st.session_state.cfs_dictionary_data, label)
                                    if 'REFERENCE' in label_filtered.columns:
                                        unique_mappings = label_filtered[['REFERENCE', 'MNEMONIC']].drop_duplicates().sort_values('REFERENCE')
                                        manual_options = ['None'] + [f"{row['REFERENCE']} → {row['MNEMONIC']}" for _, row in unique_mappings.iterrows()]
                                    else:
                                        unique_mappings = label_filtered[['ACCOUNT', 'MNEMONIC']].drop_duplicates().sort_values('ACCOUNT')
                                        manual_options = ['None'] + [f"{row['ACCOUNT']} → {row['MNEMONIC']}" for _, row in unique_mappings.iterrows()]
                                    manual_selection = st.selectbox(
                                        "Select manually",
                                        options=manual_options,
                                        key=f"manual{idx}",
                                        label_visibility="collapsed"
                                    )
                                with cols[3]:
                                    st.markdown("✅ Final Choice")
                                    choice_options = ['Select...']
                                    if matches['fuzzy']:
                                        choice_options.append(f"Fuzzy ({matches['fuzzy']['score']}%)")
                                    if matches['llm'] and 'error' not in matches['llm']:
                                        choice_options.append("AI Match")
                                    choice_options.append("Manual Override")
                                    choice_options.append("Leave Unmapped")
                                    choice_options.append("Skip/Remove")
                                    default_choice = 0
                                    if idx in st.session_state.cfs_user_selections:
                                        prev_choice = st.session_state.cfs_user_selections[idx]['choice']
                                        if prev_choice in choice_options:
                                            default_choice = choice_options.index(prev_choice)
                                    else:
                                        if matches['fuzzy'] and matches['fuzzy']['score'] >= high_confidence_threshold:
                                            default_choice = 1
                                        elif matches['llm'] and 'error' not in matches['llm']:
                                            default_choice = 2 if "AI Match" in choice_options else default_choice
                                        elif matches['fuzzy'] and matches['fuzzy']['score'] >= fuzzy_threshold:
                                            default_choice = 1
                                    final_choice = st.selectbox(
                                        "Choose mapping",
                                        options=choice_options,
                                        key=f"choice_{idx}",
                                        index=default_choice,
                                        label_visibility="collapsed"
                                    )
                                    # Bug 3 Fix: Only update state if the selection changed to prevent screen jitter
                                    current_selection = st.session_state.cfs_user_selections.get(idx, {})
                                    new_manual = manual_selection if manual_selection != 'None' else None
                                    if (idx not in st.session_state.cfs_user_selections or
                                        current_selection.get('choice') != final_choice or
                                        current_selection.get('manual') != new_manual):
                                        st.session_state.cfs_user_selections[idx] = {
                                            'choice': final_choice,
                                            'manual': new_manual,
                                            'fuzzy': matches['fuzzy'],
                                            'llm': matches['llm']
                                        }
                                st.divider()
                st.markdown("### Mapping Summary")
                total_items = len(st.session_state.cfs_mapping_results)
                fuzzy_high = sum(1 for r in st.session_state.cfs_mapping_results if r['matches']['fuzzy'] and r['matches']['fuzzy']['score'] >= high_confidence_threshold)
                fuzzy_medium = sum(1 for r in st.session_state.cfs_mapping_results if r['matches']['fuzzy'] and fuzzy_threshold <= r['matches']['fuzzy']['score'] < high_confidence_threshold)
                fuzzy_low = sum(1 for r in st.session_state.cfs_mapping_results if r['matches']['fuzzy'] and r['matches']['fuzzy']['score'] < fuzzy_threshold)
                mapped_count = sum(1 for sel in st.session_state.cfs_user_selections.values() if sel['choice'] not in ['Select...', '', 'Skip/Remove', 'Leave Unmapped'])
                unmapped_count = sum(1 for sel in st.session_state.cfs_user_selections.values() if sel['choice'] == 'Leave Unmapped')
                skip_count = sum(1 for sel in st.session_state.cfs_user_selections.values() if sel['choice'] == 'Skip/Remove')
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Items", total_items)
                with col2:
                    st.metric(f"High Confidence (≥{high_confidence_threshold}%)", f"{fuzzy_high} ({fuzzy_high/total_items*100:.0f}%)")
                with col3:
                    st.metric(f"Medium Confidence ({fuzzy_threshold}-{high_confidence_threshold-1}%)", f"{fuzzy_medium} ({fuzzy_medium/total_items*100:.0f}%)")
                with col4:
                    st.metric(f"Below Threshold (<{fuzzy_threshold}%)", f"{fuzzy_low} ({fuzzy_low/total_items*100:.0f}%)")
                st.markdown("#### Selection Summary")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Mapped", mapped_count, f"{mapped_count/total_items*100:.0f}%")
                with col2:
                    st.metric("Unmapped", unmapped_count, help="Items deliberately left unmapped")
                with col3:
                    st.metric("Skipped", skip_count)
                with col4:
                    pending = total_items - mapped_count - unmapped_count - skip_count
                    st.metric("Pending", pending)
                if st.session_state.auto_update_dict and not st.session_state.show_success_message:
                    st.divider()
                    st.markdown("### 📝 Update Data Dictionary")
                    with st.expander("ℹ️ How Dictionary Updates Work", expanded=False):
                        st.markdown("""
What gets added:
* New account names that don't exist in the dictionary
* The reference account they're mapped to
* The mnemonic (standard code) for reporting
Benefits:
* Future files with these account names will match automatically
* Improves fuzzy matching accuracy over time
* Builds company-specific knowledge base
""")
                    new_mappings = []
                    if st.session_state.cfs_user_selections and st.session_state.cfs_mapping_results:
                        for idx, selection in st.session_state.cfs_user_selections.items():
                            if selection['choice'] not in ['Select...', '', 'Skip/Remove', 'Leave Unmapped']:
                                result = next(r for r in st.session_state.cfs_mapping_results if r['index'] == idx)
                                account_name = result['account']
                                label = result['label']
                                existing = st.session_state.cfs_dictionary_data[
                                    (st.session_state.cfs_dictionary_data['ACCOUNT'] == account_name) &
                                    (st.session_state.cfs_dictionary_data['LABEL'] == label)
                                ]
                                if existing.empty:
                                    if selection['choice'].startswith('Fuzzy'):
                                        mnemonic = selection['fuzzy']['mnemonic']
                                        reference = selection['fuzzy']['match']
                                    elif selection['choice'] == 'AI Match':
                                        mnemonic = selection['llm']['mnemonic']
                                        reference = selection['llm']['match']
                                    elif selection['choice'] == 'Manual Override' and selection['manual']:
                                        if ' → ' in selection['manual']:
                                            reference, mnemonic = selection['manual'].split(' → ')
                                        else:
                                            continue
                                    else:
                                        continue
                                    if account_name and label and mnemonic and reference:
                                        new_mappings.append({
                                            'ACCOUNT': account_name,
                                            'LABEL': label,
                                            'MNEMONIC': mnemonic,
                                            'REFERENCE': reference
                                        })
                    if new_mappings:
                        st.info(f"🆕 Found {len(new_mappings)} new account mapping(s) not in the dictionary")
                        st.markdown("#### New Mappings to Add:")
                        st.caption("These account names will be added to the dictionary with their mapped references")
                        for i, mapping in enumerate(new_mappings):
                            with st.expander(f"{mapping['ACCOUNT']} → {mapping['MNEMONIC']}", expanded=True):
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.write(f"Account: {mapping['ACCOUNT']}")
                                    st.write(f"Label: {mapping['LABEL']}")
                                with col2:
                                    st.write(f"Reference: {mapping['REFERENCE']}")
                                    st.write(f"Mnemonic: {mapping['MNEMONIC']}")
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            if st.button("➕ Add All to Dictionary", type="primary", key="add_all_to_dict"):
                                try:
                                    success_count = 0
                                    added_mappings = []
                                    for mapping in new_mappings:
                                        insert_query = f"""
                                            INSERT INTO LTMA.PUBLIC.CASH_FLOW_STATEMENT_COMPLEX_DICTIONARY
                                            (ACCOUNT, LABEL, MNEMONIC, REFERENCE)
                                            VALUES (
                                                '{mapping['ACCOUNT'].replace("'", "''")}',
                                                '{mapping['LABEL'].replace("'", "''")}',
                                                '{mapping['MNEMONIC'].replace("'", "''")}',
                                                '{mapping['REFERENCE'].replace("'", "''")}'
                                            )
                                        """
                                        try:
                                            session.sql(insert_query).collect()
                                            success_count += 1
                                            added_mappings.append(mapping)
                                        except Exception as e:
                                            st.warning(f"Failed to add {mapping['ACCOUNT']}: {str(e)}")
                                    if success_count > 0:
                                        # Clear cache and reload
                                        load_dictionary_cached.clear()
                                        dict_df = load_dictionary_cached(cache_key)
                                        st.session_state.cfs_dictionary_data = dict_df
                                        st.session_state.just_added_mappings = added_mappings
                                        st.session_state.show_success_message = True
                                        st.session_state.cfs_mapping_results = None
                                        st.rerun()
                                except Exception as e:
                                    st.error(f"Error adding to dictionary: {str(e)}")
                        with col2:
                            new_df = pd.DataFrame(new_mappings)
                            csv = new_df.to_csv(index=False)
                            st.download_button(
                                "📥 Download New Mappings",
                                csv,
                                "new_cash_flow_mappings.csv",
                                "text/csv"
                            )
                        with col3:
                            st.info("💡 Adding to dictionary improves future matching")
                    else:
                        st.success("✅ No new mappings to add - all selected accounts already exist in dictionary")
                st.divider()
                if st.button("📋 Generate Final Output", type="primary", use_container_width=True, key="generate_final_output"):
                    if st.session_state.cfs_user_selections:
                        final_df = aggregated.copy()
                        final_df['Selected_Method'] = ''
                        final_df['Matched_Account'] = ''
                        final_df['Mnemonic'] = ''
                        final_df['Reference'] = ''
                        for idx, selection in st.session_state.cfs_user_selections.items():
                            if selection['choice'] == 'Skip/Remove':
                                final_df.at[idx, 'Selected_Method'] = 'REMOVE'
                            elif selection['choice'].startswith('Fuzzy'):
                                final_df.at[idx, 'Selected_Method'] = 'Fuzzy'
                                final_df.at[idx, 'Matched_Account'] = selection['fuzzy']['match']
                                final_df.at[idx, 'Mnemonic'] = selection['fuzzy']['mnemonic']
                                final_df.at[idx, 'Reference'] = selection['fuzzy']['match']
                            elif selection['choice'] == 'AI Match':
                                final_df.at[idx, 'Selected_Method'] = 'AI'
                                final_df.at[idx, 'Matched_Account'] = selection['llm']['match']
                                final_df.at[idx, 'Mnemonic'] = selection['llm']['mnemonic']
                                final_df.at[idx, 'Reference'] = selection['llm']['match']
                            elif selection['choice'] == 'Manual Override':
                                final_df.at[idx, 'Selected_Method'] = 'Manual'
                                if selection['manual'] and ' → ' in selection['manual']:
                                    reference, mnemonic = selection['manual'].split(' → ')
                                    final_df.at[idx, 'Matched_Account'] = reference
                                    final_df.at[idx, 'Mnemonic'] = mnemonic
                                    final_df.at[idx, 'Reference'] = reference
                            elif selection['choice'] == 'Leave Unmapped':
                                final_df.at[idx, 'Selected_Method'] = 'Unmapped'
                                final_df.at[idx, 'Mnemonic'] = ''
                                final_df.at[idx, 'Reference'] = final_df.at[idx, 'Account']
                        final_df = final_df[final_df['Selected_Method'] != 'REMOVE']
                        numeric_cols = [col for col in final_df.columns if col not in ['Label', 'Account', 'Selected_Method', 'Matched_Account', 'Mnemonic', 'Reference']]
                        standardized_data = final_df[final_df['Selected_Method'] != 'Unmapped']
                        if not standardized_data.empty:
                            # Group by Label, Reference, and Mnemonic for standardized sheet
                            standardized = standardized_data.groupby(['Label', 'Reference', 'Mnemonic'], as_index=False)[numeric_cols].sum()
                            # Rename columns as requested
                            standardized = standardized.rename(columns={
                                'Reference': 'Final Mnemonic Selection',
                                'Mnemonic': 'CIQ'
                            })
                            # Reorder columns for Standardized Sheet
                            column_order = ['Label', 'Final Mnemonic Selection', 'CIQ'] + numeric_cols
                            standardized = standardized[column_order]
                            # Sort by label order
                            standardized = sort_by_label_and_account(standardized, sort_column='Final Mnemonic Selection')
                        else:
                            standardized = pd.DataFrame()
                        # Create the 'As Presented' sheet with the specified columns
                        as_presented = final_df[['Label', 'Account', 'Reference'] + numeric_cols]
                        as_presented = as_presented.rename(columns={'Reference': 'Final Mnemonic Selection'})
                        as_presented = sort_by_label_and_account(as_presented) # Sorts by 'Account' by default
                        company_clean = re.sub(r'[^\w\s-]', '', st.session_state.company_name).strip()
                        company_clean = re.sub(r'[-\s]+', '_', company_clean)
                        filename = f"{company_clean}_mapped_cash_flow.xlsx" if company_clean else "mapped_cash_flow.xlsx"
                        excel_file = io.BytesIO()
                        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                            # Cover sheet
                            cover_data = {
                                'Field': ['Company', 'Currency', 'Magnitude', 'Processing Date', 'Total Mapped Items', 'Total Unmapped Items'],
                                'Value': [st.session_state.company_name, currency, magnitude, datetime.now().strftime('%Y-%m-%d'), mapped_count, unmapped_count]
                            }
                            pd.DataFrame(cover_data).to_excel(writer, sheet_name='Cover', index=False)
                            # Standardized sheet with updated name and columns
                            if not standardized.empty:
                                standardized.to_excel(writer, sheet_name='Standardized - Cash Flow', index=False)
                            # As Presented sheet with updated name
                            as_presented.to_excel(writer, sheet_name='As Presented - Cash Flow', index=False)
                            # Mapping Details sheet
                            mapping_details = final_df[['Account', 'Label', 'Selected_Method', 'Matched_Account', 'Mnemonic']]
                            mapping_details.to_excel(writer, sheet_name='Mapping Details', index=False)
                        excel_file.seek(0)
                        st.success(f"✅ Output generated for {st.session_state.company_name}!")
                        st.download_button(
                            f"📥 Download {filename}",
                            excel_file,
                            filename,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.warning("No mappings selected. Please select mappings first.")
        elif st.session_state.workflow_step == "📚 Data Dictionary Management":
            st.markdown('<div class="section-header">Data Dictionary Management</div>', unsafe_allow_html=True)
            st.markdown("<p>View, manage, and export the account mapping dictionary</p>", unsafe_allow_html=True)
            # Auto-load dictionary if not loaded
            if st.session_state.cfs_dictionary_data is None:
                with st.spinner("Loading dictionary..."):
                    dict_data = load_dictionary_cached(cache_key)
                    if dict_data is not None:
                        st.session_state.cfs_dictionary_data = dict_data
                    else:
                        st.error("Failed to load dictionary")
            if st.button("🔄 Refresh Dictionary", use_container_width=True, key="refresh_dictionary_btn"):
                with st.spinner("Reloading dictionary from database..."):
                    load_dictionary_cached.clear()
                    dict_data = load_dictionary_cached(cache_key)
                    st.session_state.cfs_dictionary_data = dict_data
                    if dict_data is not None:
                        st.success(f"Refreshed! Loaded {len(dict_data)} entries")
                    else:
                        st.error("Failed to refresh dictionary.")
                st.rerun()
            if st.session_state.cfs_dictionary_data is not None:
                dict_df = st.session_state.cfs_dictionary_data
                st.markdown("#### Dictionary Statistics")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Entries", len(dict_df))
                with col2:
                    st.metric("Unique References", dict_df['REFERENCE'].nunique())
                with col3:
                    st.metric("Unique Mnemonics", dict_df['MNEMONIC'].nunique())
                with col4:
                    st.metric("Labels", dict_df['LABEL'].nunique())
                st.markdown("#### View Dictionary")
                view_mode = st.radio("View Mode", ["Standard View", "Reference Grouping", "Remove Records"], horizontal=True, key="dict_view_mode")
                if view_mode == "Reference Grouping":
                    st.markdown("##### Account Variations by Reference/Mnemonic")
                    groupby_cols = ['LABEL', 'REFERENCE', 'MNEMONIC']
                    reference_groups = dict_df.groupby(groupby_cols).agg(
                        Variation_Count=('ACCOUNT', 'count'),
                        Account_Variations=('ACCOUNT', lambda x: list(x))
                    ).reset_index()
                    label_filter = st.selectbox(
                        "Filter by Label",
                        ['All'] + sorted(dict_df['LABEL'].unique()),
                        key="dict_ref_label_filter"
                    )
                    if label_filter != 'All':
                        reference_groups = reference_groups[reference_groups['LABEL'] == label_filter]
                    multi_variations = reference_groups[reference_groups['Variation_Count'] > 1].sort_values('Variation_Count', ascending=False)
                    if not multi_variations.empty:
                        st.markdown(f"**References with multiple account variations ({len(multi_variations)} total)**")
                        for _, row in multi_variations.head(25).iterrows():
                            with st.expander(f"`{row['REFERENCE']}` ({row['Variation_Count']} variations)"):
                                st.write(f"**Label**: {row['LABEL']}")
                                st.write(f"**Mnemonic**: {row['MNEMONIC']}")
                                st.write("**Account Variations**:")
                                for acc in row['Account_Variations']:
                                    st.write(f"- `{acc}`")
                    else:
                        st.info("No references have multiple account variations for the selected filter.")
                elif view_mode == "Remove Records":
                    st.markdown("##### Remove Dictionary Records")
                    st.warning("⚠️ **Caution**: Removing records will permanently delete them from the dictionary. This action cannot be undone.", icon="🔥")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        rm_label_filter = st.selectbox(
                            "Filter by Label",
                            ['All'] + sorted(dict_df['LABEL'].unique()),
                            key="rm_label_filter"
                        )
                    with col2:
                        rm_search = st.text_input("Search Account name", key="rm_search")
                    with col3:
                        rm_mnemonic = st.text_input("Search Mnemonic", key="rm_mnemonic")
                    filtered_remove = dict_df.copy()
                    if rm_label_filter != 'All':
                        filtered_remove = filtered_remove[filtered_remove['LABEL'] == rm_label_filter]
                    if rm_search:
                        filtered_remove = filtered_remove[filtered_remove['ACCOUNT'].str.contains(rm_search, case=False, na=False)]
                    if rm_mnemonic:
                        filtered_remove = filtered_remove[filtered_remove['MNEMONIC'].str.contains(rm_mnemonic, case=False, na=False)]
                    if not filtered_remove.empty:
                        # Create a unique identifier for multiselect
                        filtered_remove['display_id'] = filtered_remove.apply(
                            lambda row: f"ACCOUNT: {row['ACCOUNT']} | LABEL: {row['LABEL']} | MNEMONIC: {row['MNEMONIC']}", axis=1
                        )
                        options = filtered_remove['display_id'].tolist()
                        records_to_remove_ids = st.multiselect(
                            "Select records to remove:",
                            options=options,
                            key="remove_multiselect"
                        )
                        if records_to_remove_ids:
                            records_to_remove_df = filtered_remove[filtered_remove['display_id'].isin(records_to_remove_ids)]
                            selected_count = len(records_to_remove_df)
                            st.error(f"You have selected {selected_count} record(s) for permanent deletion.", icon="🗑️")
                            st.dataframe(records_to_remove_df[['ACCOUNT', 'LABEL', 'MNEMONIC', 'REFERENCE']])
                            if st.button(f"Confirm and Remove {selected_count} Record(s)", type="primary"):
                                with st.spinner("Removing records..."):
                                    success_count = 0
                                    errors = []
                                    def build_where_condition(column_name, value):
                                        col = f'"{column_name}"'
                                        if pd.isna(value) or value is None:
                                            return f"{col} IS NULL"
                                        else:
                                            escaped_value = str(value).replace("'", "''")
                                            return f"{col} = '{escaped_value}'"
                                    for _, row in records_to_remove_df.iterrows():
                                        try:
                                            key_columns = ['ACCOUNT', 'LABEL', 'MNEMONIC', 'REFERENCE']
                                            where_conditions = [build_where_condition(col, row.get(col)) for col in key_columns]
                                            where_clause = " AND ".join(where_conditions)
                                            delete_query = f"DELETE FROM LTMA.PUBLIC.CASH_FLOW_STATEMENT_COMPLEX_DICTIONARY WHERE {where_clause}"
                                            result = session.sql(delete_query).collect()
                                            if result[0]['number of rows deleted'] > 0:
                                                success_count += 1
                                            else:
                                                errors.append(f"No match found in DB for Account '{row['ACCOUNT']}'. It may have already been deleted.")
                                        except Exception as e:
                                            errors.append(f"Failed to remove Account '{row['ACCOUNT']}': {str(e)}")
                                    st.success(f"✅ Successfully removed {success_count} of {selected_count} selected record(s).")
                                    if errors:
                                        st.error("Some errors occurred during removal:")
                                        for error in errors:
                                            st.write(error)
                                    load_dictionary_cached.clear()
                                    st.session_state.cfs_dictionary_data = None
                                    st.rerun()
                    else:
                        st.info("No records found matching your filters.")
                else: # Standard View
                    col1, col2 = st.columns(2)
                    with col1:
                        label_filter = st.selectbox(
                            "Filter by Label",
                            ['All'] + sorted(dict_df['LABEL'].unique()),
                            key="dict_std_label_filter"
                        )
                    with col2:
                        search = st.text_input("Search (Account/Reference/Mnemonic)", key="dict_std_search")
                    filtered = dict_df.copy()
                    if label_filter != 'All':
                        filtered = filtered[filtered['LABEL'] == label_filter]
                    if search:
                        search_mask = (
                            filtered['ACCOUNT'].str.contains(search, case=False, na=False) |
                            filtered['REFERENCE'].str.contains(search, case=False, na=False) |
                            filtered['MNEMONIC'].str.contains(search, case=False, na=False)
                        )
                        filtered = filtered[search_mask]
                    st.markdown("---")
                    ITEMS_PER_PAGE_DICT = 50
                    total_items = len(filtered)
                    total_pages = (total_items // ITEMS_PER_PAGE_DICT) + (1 if total_items % ITEMS_PER_PAGE_DICT > 0 else 0)
                    total_pages = max(total_pages, 1)
                    if st.session_state.dict_page_number > total_pages:
                        st.session_state.dict_page_number = total_pages
                    page_col1, page_col2, page_col3 = st.columns([1, 2, 1])
                    with page_col1:
                        if st.button("⬅️ Previous", key="dict_prev") and st.session_state.dict_page_number > 1:
                            st.session_state.dict_page_number -= 1
                            st.rerun()
                    with page_col2:
                        page_num_str = f"Page {st.session_state.dict_page_number} of {total_pages}"
                        st.write(f'<div style="text-align: center;">{page_num_str}</div>', unsafe_allow_html=True)
                    with page_col3:
                        if st.button("Next ➡️", key="dict_next") and st.session_state.dict_page_number < total_pages:
                            st.session_state.dict_page_number += 1
                            st.rerun()
                    start_idx = (st.session_state.dict_page_number - 1) * ITEMS_PER_PAGE_DICT
                    end_idx = start_idx + ITEMS_PER_PAGE_DICT
                    display_cols = ['ACCOUNT', 'LABEL', 'MNEMONIC', 'REFERENCE']
                    st.dataframe(filtered.iloc[start_idx:end_idx][display_cols], use_container_width=True, height=500)
                st.markdown("#### Export Dictionary")
                col1, col2, col3 = st.columns(3)
                with col1:
                    csv_full = dict_df.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        "📥 Export Full Dictionary",
                        data=csv_full,
                        file_name="cash_flow_dictionary_full.csv",
                        mime="text/csv",
                        key="export_full_dict",
                        use_container_width=True
                    )
                with col2:
                    ref_list = dict_df[['LABEL', 'REFERENCE', 'MNEMONIC']].drop_duplicates()
                    csv_ref = ref_list.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        "📥 Export Reference List",
                        data=csv_ref,
                        file_name="cash_flow_references.csv",
                        mime="text/csv",
                        key="export_ref_list",
                        use_container_width=True
                    )
                with col3:
                    training_data = dict_df.copy()
                    training_data['Mapping'] = training_data['ACCOUNT'] + ' → ' + training_data['REFERENCE']
                    csv_train = training_data[['LABEL', 'Mapping', 'MNEMONIC']].to_csv(index=False).encode('utf-8')
                    st.download_button(
                        "📥 Export for Training",
                        data=csv_train,
                        file_name="cash_flow_training_data.csv",
                        mime="text/csv",
                        key="export_training",
                        use_container_width=True
                    )
if __name__ == "__main__":
    main()